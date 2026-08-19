"""
Employee directory.

An admin uploads the client's employee list once (Admin → Settings → Employee
directory). Every employee belongs to exactly one GROUP — in practice the Cost
Center column of the client's HR export, though the app doesn't care what the
source column is called. Initiatives are stamped with a group, and the people
who can be named on an initiative are the people in that group.

Two things this deliberately does NOT do:

  * It does not touch the sign-in list. EMPLOYEES in app.py stays as the set of
    people who log in and prepare reports. The directory is who can be named as
    a team member — 2,000+ names for the current client, unusable as a sign-in
    dropdown.

  * It does not filter by employment status. Someone who left in September still
    did R&D work in Q1, so terminated employees stay selectable and unmarked.

Storage is a single JSON file alongside the reports. On ephemeral hosting it is
lost on restart along with everything else; re-uploading takes seconds.
"""

from __future__ import annotations

import io
import json
import time
from pathlib import Path

from openpyxl import load_workbook

DATA_DIR = Path("data")
DIR_FILE = DATA_DIR / "employee_directory.json"

# Header names recognised automatically when mapping an uploaded file.
# Order matters — earlier is preferred.
NAME_HINTS = (
    "preferred name", "employee name", "full name", "legal name",
    "worker name", "name",
)
GROUP_HINTS = (
    "cost center", "cost centre", "costcenter", "cost_center",
    "sub function", "department", "group",
)
ID_HINTS = (
    "employee id", "employee number", "worker id", "emp id", "employee #",
)

NO_GROUP = "(no group)"

_cache: dict = {"stamp": None, "data": None}


# ── Reading an uploaded workbook ─────────────────────────────────────────────

def sheet_names(file_bytes: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _hint_score(text) -> int:
    t = str(text).strip().lower()
    if not t:
        return 0
    for hints in (NAME_HINTS, GROUP_HINTS, ID_HINTS):
        if any(t == h for h in hints):
            return 2
        if any(h in t for h in hints):
            return 1
    return 0


def find_header(rows: list[list], max_scan: int = 30) -> int:
    """Index (0-based) of the row holding the column names.

    HR exports routinely carry a title, a note, and blank rows above the real
    header — the current client's file has five. Scoring on recognised header
    names rather than assuming row 1 is what lets those files load without the
    client having to clean them up first.
    """
    best_i, best_score = -1, 0.0
    for i, row in enumerate(rows[:max_scan]):
        filled = sum(1 for c in row if str(c or "").strip())
        if filled < 2:
            continue
        score = sum(_hint_score(c) for c in row) + min(filled, 10) * 0.05
        if score > best_score:
            best_i, best_score = i, score
    if best_i >= 0 and best_score >= 1:
        return best_i
    for i, row in enumerate(rows[:max_scan]):
        if sum(1 for c in row if str(c or "").strip()) >= 2:
            return i
    return 0


def _read_rows(file_bytes: bytes, sheet: str) -> list[list]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()


def _labels(header: list) -> list:
    """Column labels, duplicates disambiguated so a pick maps back to exactly
    one index. None for empty header cells."""
    out = []
    seen: dict = {}
    for cell in header:
        label = str(cell).strip() if cell is not None else ""
        if not label:
            out.append(None)
            continue
        if label in seen:
            seen[label] += 1
            label = f"{label} ({seen[label]})"
        else:
            seen[label] = 1
        out.append(label)
    return out


def _suggest(columns: list, hints: tuple) -> str:
    lowered = [(c, c.strip().lower()) for c in columns]
    for hint in hints:
        for original, low in lowered:
            if low == hint:
                return original
    for hint in hints:
        for original, low in lowered:
            if hint in low:
                return original
    return ""


def inspect(file_bytes: bytes, sheet: str) -> dict:
    """What the admin screen needs to render the column-mapping controls."""
    rows = _read_rows(file_bytes, sheet)
    if not rows:
        return {"header_row": 0, "columns": [], "suggested": {}, "n_rows": 0}

    hi = find_header(rows)
    columns = [c for c in _labels(rows[hi]) if c]
    n_rows = sum(1 for r in rows[hi + 1:] if any(str(c or "").strip() for c in r))

    return {
        "header_row": hi + 1,          # 1-based, for display
        "columns": columns,
        "suggested": {
            "name":  _suggest(columns, NAME_HINTS),
            "group": _suggest(columns, GROUP_HINTS),
            "id":    _suggest(columns, ID_HINTS),
        },
        "n_rows": n_rows,
    }


def parse(
    file_bytes: bytes,
    sheet: str,
    name_col: str,
    group_col: str,
    id_col: str = "",
):
    """Turn the chosen sheet into directory records.

    Returns (records, warnings). Each record is {"name", "group", "id",
    "display"}, where `display` is what pickers show and what gets stored on an
    initiative's team_members.
    """
    rows = _read_rows(file_bytes, sheet)
    if not rows:
        return [], ["That sheet is empty."]

    hi = find_header(rows)
    labels = _labels(rows[hi])

    def index_of(label: str) -> int:
        return labels.index(label) if label in labels else -1

    ni, gi = index_of(name_col), index_of(group_col)
    ii = index_of(id_col) if id_col else -1
    if ni < 0:
        return [], [f"Couldn't find the name column '{name_col}' on this sheet."]
    if gi < 0:
        return [], [f"Couldn't find the group column '{group_col}' on this sheet."]

    def cell(row, idx):
        if idx < 0 or idx >= len(row):
            return ""
        v = row[idx]
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return str(v).strip()

    records = []
    blank_names = blank_groups = 0

    for row in rows[hi + 1:]:
        name = cell(row, ni)
        if not name:
            if any(str(c or "").strip() for c in row):
                blank_names += 1
            continue
        group = cell(row, gi) or NO_GROUP
        if group == NO_GROUP:
            blank_groups += 1
        records.append({
            "name": name, "group": group, "id": cell(row, ii), "display": name,
        })

    if not records:
        return [], ["No employees found — check the sheet and the name column."]

    # Duplicate names are ambiguous once sitting in a team_members list, so
    # append the employee ID to any name that isn't unique.
    counts_by_name: dict = {}
    for r in records:
        counts_by_name[r["name"]] = counts_by_name.get(r["name"], 0) + 1
    dupes = sorted(n for n, c in counts_by_name.items() if c > 1)
    for r in records:
        if counts_by_name[r["name"]] > 1:
            r["display"] = (
                f"{r['name']} ({r['id']})" if r["id"] else f"{r['name']} — {r['group']}"
            )

    warnings = []
    if blank_names:
        warnings.append(
            f"{blank_names} row{'s' if blank_names != 1 else ''} had no name and "
            "were skipped."
        )
    if blank_groups:
        warnings.append(
            f"{blank_groups} employee{'s' if blank_groups != 1 else ''} had no "
            f"group — they're collected under '{NO_GROUP}'."
        )
    if dupes:
        shown = ", ".join(dupes[:3]) + ("…" if len(dupes) > 3 else "")
        warnings.append(
            f"{len(dupes)} duplicate name{'s' if len(dupes) != 1 else ''} ({shown}) — "
            "the employee ID was appended to tell them apart."
        )

    return records, warnings


# ── Persistence ───────────────────────────────────────────────────────────────

def save(records: list, meta_info: dict) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    payload = {
        "meta": {**meta_info, "saved_at": int(time.time() * 1000),
                 "n_employees": len(records)},
        "records": records,
    }
    tmp = DIR_FILE.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload), encoding="utf-8")
    tmp.replace(DIR_FILE)
    _cache["stamp"] = None


def load() -> dict:
    """{"meta", "records", "_by_group", "_index"} — empty if nothing loaded.

    Cached on the file's mtime and size, so a 2,000-row directory is read and
    indexed once rather than on every Streamlit rerun.
    """
    empty = {"meta": {}, "records": [], "_by_group": {}, "_index": {}}
    if not DIR_FILE.exists():
        _cache["stamp"], _cache["data"] = None, None
        return empty

    stat = DIR_FILE.stat()
    stamp = (stat.st_mtime_ns, stat.st_size)
    if _cache["stamp"] == stamp and _cache["data"] is not None:
        return _cache["data"]

    try:
        data = json.loads(DIR_FILE.read_text(encoding="utf-8"))
    except Exception:
        return empty

    data.setdefault("meta", {})
    records = data.setdefault("records", [])

    by_group: dict = {}
    index: dict = {}
    for r in records:
        # "cc" is what an earlier build wrote; read either field name.
        group = r.get("group") or r.get("cc") or NO_GROUP
        r["group"] = group
        display = r.get("display") or r.get("name", "")
        by_group.setdefault(group, []).append(display)
        index[display] = group
    for names in by_group.values():
        names.sort(key=str.casefold)

    data["_by_group"], data["_index"] = by_group, index
    _cache["stamp"], _cache["data"] = stamp, data
    return data


def clear() -> bool:
    _cache["stamp"], _cache["data"] = None, None
    if DIR_FILE.exists():
        DIR_FILE.unlink()
        return True
    return False


def is_loaded() -> bool:
    return bool(load()["records"])


def meta() -> dict:
    return load()["meta"]


# ── Lookups ───────────────────────────────────────────────────────────────────

def groups() -> list:
    """Every group, alphabetical.

    The full source value is kept, code prefix and all. That is load-bearing:
    the current client reuses a name across two different codes — "GCO,
    Americas" is both 1300H00025 (104 people) and 1300I00024 (53) — so
    stripping the code would silently merge two different groups.
    """
    return sorted(load()["_by_group"].keys(), key=str.casefold)


def names_in(group: str) -> list:
    return list(load()["_by_group"].get(group, []))


def all_names() -> list:
    return sorted(load()["_index"].keys(), key=str.casefold)


def group_of(display_name: str) -> str:
    return load()["_index"].get(display_name, "")


def known(display_name: str) -> bool:
    """False for write-ins — contractors and anyone missing from the file."""
    return display_name in load()["_index"]


def counts() -> tuple:
    """(employees, groups)"""
    d = load()
    return len(d["records"]), len(d["_by_group"])
