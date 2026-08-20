"""
R&D Monthly Activity Tracker
Run with: streamlit run app.py
"""

import streamlit as st
import json
from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re
import time

import store
import directory

# ── Config ────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="R&D Activity Tracker",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Constants ─────────────────────────────────────────────────────────────────

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

# Who can SIGN IN and prepare reports. Deliberately NOT the list of people who
# can be named on an initiative — that comes from the uploaded employee
# directory (directory.py), which runs to thousands of names for a real client
# and would be unusable as a sign-in dropdown.
#
# EMPLOYEES is also the fallback for team selection until a directory is
# uploaded, so a fresh install still works.
EMPLOYEES = [
    "Bob Smith", "Sara", "Doug", "Trevor", "Doni",
    "Jonathan", "Steven", "Michael", "Joe", "Nicole Browne",
]

# What the app calls the grouping that employees are filtered by. It maps to the
# client's Cost Center column, but "Group" alone would collide with the Group
# number in export filenames (107_Group_10_2025_...), which means the month.
# Change this one string to relabel it everywhere, including the export header.
GROUP_LABEL = "Employee Group"

ENTITIES = ["107", "108", "109", "110"]   # base/default entities — custom ones persist separately

STATUS_LABELS = {
    "in-progress":       "🟡 In Progress",
    "submitted":         "🔵 Ready for Review",
    "rejected":          "🔴 Rejected",
    "approved":          "🟢 Approved",
    "archived":          "📦 Archived",
    "not-started":       "⚪ Not Started",
}

# Report status is DERIVED from the initiatives it contains rather than set
# directly. Previously, returning a single initiative flipped the whole report to
# "rejected", locked the user out of editing until they clicked an
# acknowledgement button, and left a report reading as rejected while most of its
# initiatives were approved. See computed_status().
#
# A report in "rejected" stays editable — that is the whole point.
LOCKED_STATUSES = ("submitted", "approved", "archived")

# Kept for the import parser, which may meet labels from older exports.
LEGACY_STATUS_ALIASES = {"changes requested": "rejected"}

# Wizard steps — exactly the 9 columns in the Excel template
# (Month/Yr is derived from reporting_month, not asked separately)
WIZARD_STEPS = [
    {
        # Asked first, because it determines who can be named on this
        # initiative. Changing it later clears any team member who isn't in the
        # new group — see the group branch in screen_wizard().
        "field": "employee_group",
        "label": GROUP_LABEL,
        "type": "group",
        "question": f"Which {GROUP_LABEL.lower()} is this initiative's work in?",
        "hint": "Team members you pick later are limited to this group. "
                "Type to search — the code and the name both match.",
        "required": True,
    },
    {
        "field": "business_component",
        "label": "Business Component",
        "type": "text",
        "placeholder": "e.g. Software Development, Hardware Engineering...",
        "question": "What business component does this initiative belong to?",
        "hint": "Enter the area of the business this R&D work supports.",
        "required": True,
    },
    {
        "field": "initiative_name",
        "label": "Initiative Name",
        "type": "text",
        "placeholder": "e.g. Automated Quality Inspection System",
        "question": "What is the name of this R&D initiative?",
        "hint": "Use a short, consistent name — this carries over month-to-month for ongoing work.",
        "required": True,
    },
    {
        "field": "initiative_description",
        "label": "Initiative Description",
        "type": "textarea",
        "placeholder": "We are developing a system that will...",
        "question": "Describe what this initiative aims to achieve.",
        "hint": "Explain the goal and expected business outcome in 2–4 sentences.",
        "required": True,
    },
    {
        "field": "tech_uncertainty",
        "label": "Tech Uncertainty",
        "type": "textarea",
        "placeholder": "It is currently unknown whether... We are testing if...",
        "question": "What technical uncertainty are you working to resolve?",
        "hint": "What scientific or technical question are you trying to answer? What don't you know yet? This is key for R&D eligibility.",
        "required": True,
    },
    {
        "field": "start_date",
        "label": "Start Date",
        "type": "date",
        "question": "When did R&D work on this initiative begin?",
        "hint": "The actual date work first started — not the project plan date.",
        "required": True,
    },
    {
        "field": "expected_end_date",
        "label": "Expected End Date",
        "type": "date",
        "question": "When do you expect this initiative to conclude?",
        "hint": "Your best estimate. This can be updated in future months.",
        "required": True,
    },
    {
        "field": "activities",
        "label": "Activities to Eliminate Technical Uncertainty",
        "type": "textarea",
        "placeholder": "Prototyping new approach, running performance tests, analyzing results...",
        "question": "What activities are being conducted to eliminate the technical uncertainty?",
        "hint": "Describe the specific tasks, experiments, or development work happening this month.",
        "required": True,
    },
    {
        "field": "team_members",
        "label": "Team Members",
        "type": "multiselect",
        "question": "Which team members are working on this initiative?",
        "hint": "Select everyone contributing to this initiative this month.",
        "required": True,
    },
    {
        "field": "notes",
        "label": "Notes",
        "type": "textarea",
        "placeholder": "Optional — blockers, scope changes, anything useful for review...",
        "question": "Any additional notes or comments?",
        "hint": "Include anything helpful for the Oversight Lead's review. This field is optional.",
        "required": False,
    },
]

CARRYOVER_FIELDS = {
    "employee_group", "contractors", "business_component", "initiative_name",
    "initiative_description", "tech_uncertainty", "start_date",
    "expected_end_date", "team_members",
}


# ── Date helpers ──────────────────────────────────────────────────────────────

def cur_month() -> str:
    return datetime.now().strftime("%Y-%m")

_ET = ZoneInfo("America/New_York")

def ts_to_et(ms: int, fmt: str = "%b %d, %Y %I:%M %p") -> str:
    """Format a millisecond UTC timestamp in US Eastern time.

    Uses a real timezone rather than a fixed offset, so daylight saving is
    handled correctly — a fixed -5 was an hour off from March to November,
    including in the export's Completion Date column.
    """
    from datetime import timezone
    if not ms:
        return ""
    dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone(_ET)
    out = dt.strftime(fmt)
    # Only label the zone when a time is actually shown.
    if any(t in fmt for t in ("%I", "%H", "%M")):
        out += " EDT" if dt.dst() else " EST"
    return out

# Kept so older call sites keep working.
ts_to_est = ts_to_et

def prev_month_of(m: str) -> str:
    y, mo = map(int, m.split("-"))
    if mo == 1:
        return f"{y-1}-12"
    return f"{y}-{str(mo-1).zfill(2)}"

def next_month_of(m: str) -> str:
    y, mo = map(int, m.split("-"))
    if mo == 12:
        return f"{y+1}-01"
    return f"{y}-{str(mo+1).zfill(2)}"

def parse_month_label(s: str) -> str:
    """Parses a display label like 'May 2026' back into '2026-05'. Returns '' if unparseable."""
    if not s:
        return ""
    s = str(s).strip()
    for fmt in ("%B %Y", "%b %Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return f"{dt.year}-{str(dt.month).zfill(2)}"
        except Exception:
            continue
    return ""

def fmt_month(m: str) -> str:
    if not m:
        return ""
    try:
        y, mo = m.split("-")
        return datetime(int(y), int(mo), 1).strftime("%B %Y")
    except Exception:
        return ""

def fmt_month_tab(m: str) -> str:
    """Short form for Excel sheet tab names, e.g. 'May 2026'. Never returns empty."""
    if not m:
        return "Report"
    try:
        y, mo = m.split("-")
        return datetime(int(y), int(mo), 1).strftime("%b %Y")
    except Exception:
        return m or "Report"

def available_months() -> list[str]:
    """6 months back + 1 forward, newest first."""
    today = date.today()
    result = []
    for delta in range(-6, 2):
        total = today.month - 1 + delta
        yr    = today.year + total // 12
        mo    = total % 12 + 1
        result.append(f"{yr}-{str(mo).zfill(2)}")
    result.sort(reverse=True)
    return result


# ── Filename helper ───────────────────────────────────────────────────────────

def export_filename(entity: str, reporting_month: str, tag: str = "") -> str:
    """
    107_Group_10_2025_Monthly_R_D_Tracking_Template.xlsx
    107_Group_10_2025_Consolidated_Submitted_Monthly_R_D_Tracking_Template.xlsx
    """
    try:
        dt = datetime.strptime(reporting_month, "%Y-%m")
        mo = str(dt.month)
        yr = str(dt.year)
    except Exception:
        mo, yr = "00", "0000"
    parts = [entity or "000", "Group", mo, yr]
    if tag:
        parts.append(tag)
    parts.append("Monthly_R_D_Tracking_Template")
    return "_".join(parts) + ".xlsx"


# ── Initiative series ─────────────────────────────────────────────────────────
# An initiative persists across months. Each month it gets its OWN ROW, stamped
# with that month, and earlier rows are kept unchanged — that is what makes a
# single file a month-by-month record of the work rather than a snapshot.
#
# series_id ties those rows together: "Initiative A in August" and
# "Initiative A in September" share a series_id but have different month_yr.

def series_id_of(init: dict) -> str:
    """Stable identity for an initiative across months.

    Falls back to a slug of the name so rows restored from an Excel export —
    which has no series column — still group correctly.
    """
    sid = (init.get("series_id") or "").strip()
    if sid:
        return sid
    name = (init.get("initiative_name") or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "-", name).strip("-") or init.get("id", "")


def is_historical(init: dict) -> bool:
    """True for a row carried in from an earlier month. Read-only."""
    return bool(init.get("historical"))


def current_rows(report: dict) -> list[dict]:
    """Rows belonging to this report's own reporting period — the editable ones."""
    return [i for i in (report.get("initiatives") or []) if not is_historical(i)]


def historical_rows(report: dict) -> list[dict]:
    """Rows carried forward from earlier months, kept for the record."""
    return [i for i in (report.get("initiatives") or []) if is_historical(i)]


def latest_per_series(report: dict) -> list[dict]:
    """The most recent row for each series in a report."""
    newest: dict[str, dict] = {}
    for init in report.get("initiatives") or []:
        sid = series_id_of(init)
        prev = newest.get(sid)
        if prev is None or (init.get("month_yr") or "") >= (prev.get("month_yr") or ""):
            newest[sid] = init
    return list(newest.values())


# ── Report status ─────────────────────────────────────────────────────────────

def computed_status(report: dict) -> str:
    """The report's status, derived from its initiatives.

    archived     set explicitly by the admin; terminal until reopened
    in-progress  never submitted
    rejected     at least one initiative sent back for revision
    approved     every initiative accepted
    submitted    waiting on review

    Only CURRENT-PERIOD rows count. Rows carried in from earlier months are a
    historical record already accepted in their own period, so they must not make
    a fresh report look approved.
    """
    if report.get("status") == "archived":
        return "archived"

    inits = report.get("initiatives") or []
    if not inits:
        return "in-progress"
    reviewable = current_rows(report) or inits
    if not report.get("submitted_at"):
        return "in-progress"
    if any(i.get("initiative_status") == "returned" for i in reviewable):
        return "rejected"
    if all(i.get("initiative_status") == "approved" for i in reviewable):
        return "approved"
    return "submitted"


def refresh_status(report: dict) -> dict:
    """Recompute and store the report's status. Called on every save."""
    report["status"] = computed_status(report)
    return report


def is_locked(report: dict) -> bool:
    """True when the preparer may not edit. rejected is NOT locked —
    the user needs to be able to fix what was flagged and resubmit."""
    return computed_status(report) in LOCKED_STATUSES


def clear_return_flag(init: dict, reason: str = "revised") -> dict:
    """Move an outstanding review comment into history and reset the initiative
    to active. Without this, a returned initiative kept showing the reviewer's
    comment as a live problem even after the user had fixed it."""
    if init.get("initiative_status") == "returned":
        comment = (init.get("review_comment") or "").strip()
        if comment:
            init.setdefault("review_history", []).append({
                "comment":     comment,
                "at":          init.get("returned_at"),
                "resolved_by": reason,
            })
        init["initiative_status"] = "active"
        init["review_comment"]    = ""
        init["returned_at"]       = None
    return init


def outstanding_returns(report: dict) -> list[dict]:
    """Current-period initiatives flagged for revision."""
    return [
        i for i in current_rows(report)
        if i.get("initiative_status") == "returned"
    ]


def review_started(report: dict) -> bool:
    """True once the reviewer has accepted or returned anything in this report.
    Until then the preparer may withdraw the submission and keep editing."""
    return any(
        i.get("initiative_status") in ("approved", "returned")
        for i in current_rows(report)
    )


def unsubmit(report: dict) -> dict:
    """Withdraw a submission so the preparer can keep working. Only valid before
    the reviewer has touched anything."""
    report.pop("submitted_at", None)
    report.pop("approved_at", None)
    return refresh_status(report)


def mark_submitted(report: dict) -> dict:
    """Submit or resubmit. Any outstanding review flags are cleared — the user is
    asserting they've addressed the feedback, and the comments stay in history so
    the reviewer can still see what was asked."""
    for init in current_rows(report):
        clear_return_flag(init, reason="resubmitted")
    report["submitted_at"] = int(time.time() * 1000)
    return refresh_status(report)


# ── Storage ───────────────────────────────────────────────────────────────────
# All persistence goes through store.py — this section is only thin adapters.
# A report is identified by (user, entity, filing period). Entity is part of the
# key, so filing two entities in the same month no longer overwrites one with
# the other.

def _safe_name(s: str) -> str:
    return s.replace(" ", "_").replace("/", "_")


# ── Entities ──────────────────────────────────────────────────────────────────

def all_entities() -> list[str]:
    """Base entities plus any added through the UI, in the order they were added."""
    custom = store.get_config("custom_entities", []) or []
    return ENTITIES + [e for e in custom if e not in ENTITIES]

def add_custom_entity(new_entity: str) -> bool:
    """Adds an entity so it persists in the dropdown. False if it already exists."""
    new_entity = str(new_entity).strip()
    if not new_entity or new_entity in all_entities():
        return False
    custom = store.get_config("custom_entities", []) or []
    custom.append(new_entity)
    store.set_config("custom_entities", custom)
    return True


# ── Business component closure ────────────────────────────────────────────────
# Closing a business component is separate from resolving a technical
# uncertainty. An uncertainty is a question that gets answered; a component is a
# body of work that finishes. One component can carry several uncertainties, so
# resolving one does not close the component, and closing the component does not
# retroactively resolve anything.
#
# Closure is stored per (entity, business component) rather than on any single
# initiative, because a component spans initiatives and months.

def _closed_bcs() -> dict:
    return store.get_config("closed_business_components", {}) or {}


def bc_is_closed(entity: str, bc: str) -> dict | None:
    """The closure record for this component, or None if it is still open."""
    key = (bc or "").strip().lower()
    if not key:
        return None
    return _closed_bcs().get(entity or "", {}).get(key)


def close_business_component(entity: str, bc: str, who: str, note: str = "") -> bool:
    """Mark a component closed. False if it was already closed."""
    key = (bc or "").strip().lower()
    if not key or bc_is_closed(entity, bc):
        return False
    data = _closed_bcs()
    data.setdefault(entity or "", {})[key] = {
        "label":     (bc or "").strip(),
        "closed_by": who,
        "closed_at": int(time.time() * 1000),
        "note":      (note or "").strip(),
    }
    store.set_config("closed_business_components", data)
    return True


def reopen_business_component(entity: str, bc: str) -> bool:
    key = (bc or "").strip().lower()
    data = _closed_bcs()
    if key in data.get(entity or "", {}):
        data[entity or ""].pop(key, None)
        store.set_config("closed_business_components", data)
        return True
    return False


def closed_bc_list(entity: str) -> list[dict]:
    """Closed components for one entity, most recently closed first."""
    recs = list(_closed_bcs().get(entity or "", {}).values())
    return sorted(recs, key=lambda r: r.get("closed_at") or 0, reverse=True)


# ── Permissions ───────────────────────────────────────────────────────────────

def load_permissions() -> dict:
    """{username: [entity, ...]}. An empty list means no restriction."""
    return store.get_config("permissions", {}) or {}

def save_permissions(perms: dict):
    store.set_config("permissions", perms)

def get_user_entities(username: str) -> list[str]:
    """Entities this user may FILE reports for.

    Note: this governs filing only. The Archive tab deliberately shows every
    entity to everyone — it is a shared record, not a per-entity permission.
    """
    perms = load_permissions()
    if username not in perms or not perms[username]:
        return all_entities()
    return [e for e in perms[username] if e in all_entities()]


# ── Reports ───────────────────────────────────────────────────────────────────

def load_submission(username: str, entity: str, month: str) -> dict | None:
    return store.get_report(username, entity, month)

def save_draft(username: str, draft: dict):
    """Persist a draft. Silently does nothing until Report Setup is complete,
    since a report with no entity or period has no identity to be stored under."""
    draft["user"] = username
    if not (draft.get("entity") and draft.get("reporting_month")):
        return
    refresh_status(draft)
    store.save_report(draft)

def all_reports(**filters) -> list[dict]:
    """Every non-empty report, newest period first. Filters pass through to store."""
    return store.list_reports(**filters)

def load_user_months(username: str) -> list[str]:
    """Filing months where this user has real data, newest first."""
    return store.list_periods(user=username)

def user_reports(username: str) -> list[dict]:
    return store.list_reports(user=username)


# ── Backup / bulk delete ──────────────────────────────────────────────────────

def create_backup_excel() -> bytes:
    """Every report, every status, one tab per entity."""
    reports = all_reports()
    if not reports:
        wb = Workbook()
        ws = wb.active
        ws.title = "No Data"
        ws.cell(1, 1, "No reports found.")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    # dedupe=False: a backup must be able to rebuild every period file, so it
    # keeps the carried-forward copies that the deliverable export collapses.
    return build_excel_consolidated(
        reports, group_by="entity", status_desc="All statuses (backup)", dedupe=False
    )


def delete_all_history() -> int:
    """Deletes every report. Entity list and permissions live in a separate
    directory and are preserved — they used to be caught by the same glob and
    wiped along with the data, with no way to recover them."""
    return store.delete_all_reports()


# ── Excel Import / Restore ─────────────────────────────────────────────────────
# Lets an admin (or user) re-upload a previously downloaded backup/export file
# to reconstruct submissions after the server's storage has been wiped
# (e.g. app restart on ephemeral hosting). Reads the same layout this app
# writes — works with the consolidated/backup format (multi-sheet, "User"
# column) and with individual single-user exports (one sheet, no "User" col).

# Reverse lookup: "Approved" -> "approved", "📦 Archived" -> "archived", etc.
_STATUS_LABEL_TO_KEY: dict[str, str] = {}
for _k, _v in STATUS_LABELS.items():
    _STATUS_LABEL_TO_KEY[_v.strip().lower()] = _k
    _parts = _v.split(" ", 1)
    if len(_parts) > 1:
        _STATUS_LABEL_TO_KEY[_parts[1].strip().lower()] = _k

def _parse_status_label(s: str) -> str:
    if not s:
        return "in-progress"
    key = _STATUS_LABEL_TO_KEY.get(str(s).strip().lower(), "")
    if not key:
        # Files exported before the status rename say "Rejected".
        key = LEGACY_STATUS_ALIASES.get(str(s).strip().lower().lstrip("🔴 ").strip(), "")
    return key or "in-progress"


def _cell_to_date_str(value) -> str:
    """Normalizes a cell value to a 'YYYY-MM-DD' string, however Excel stored it."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _find_header_row(ws, max_scan: int = 8) -> tuple[int, dict[str, int]]:
    """
    Scans the first few rows of a sheet to find the header row — identified
    as the row containing 'Initiative Name'. Returns (row_index, {lower_header: col_idx}).
    Returns (0, {}) if no header row is found.
    """
    for r in range(1, max_scan + 1):
        row_vals = {}
        found_marker = False
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            sval = str(val).strip()
            row_vals[sval.lower()] = c
            if sval.lower() == "initiative name":
                found_marker = True
        if found_marker:
            return r, row_vals
    return 0, {}


def _extract_entity(ws) -> str:
    """Tries sheet title 'Entity 107', then row 2 subtitle 'Entity: 107'."""
    m = re.search(r"Entity\s+(\w+)", ws.title or "")
    if m:
        return m.group(1)
    subtitle = ws.cell(2, 1).value or ""
    # Subtitle text may live in different columns depending on User-col offset;
    # scan the first few cells of row 2 for the pattern.
    for c in range(1, min(ws.max_column, 10) + 1):
        v = ws.cell(2, c).value
        if v and "Entity:" in str(v):
            m2 = re.search(r"Entity:\s*(\w+)", str(v))
            if m2:
                return m2.group(1)
    return ""


def _extract_filing_and_username_from_subtitle(ws) -> tuple[str, str]:
    """
    For individual (single-user) exports without a 'User' column or
    'Filing Month' column: pulls 'Submitted by: X' and 'Filing: Month Year'
    out of the row-2 subtitle text.
    """
    username, filing = "", ""
    for c in range(1, min(ws.max_column, 10) + 1):
        v = ws.cell(2, c).value
        if not v:
            continue
        sv = str(v)
        m_user = re.search(r"Submitted by:\s*([^|]+?)\s*(?:\||$)", sv)
        if m_user:
            username = m_user.group(1).strip()
        m_fil = re.search(r"Filing:\s*([A-Za-z]+ \d{4})", sv)
        if m_fil:
            filing = parse_month_label(m_fil.group(1))
    return username, filing


def parse_import_workbook(file_bytes: bytes) -> tuple[dict, list[str]]:
    """
    Parses an uploaded .xlsx (consolidated/backup or individual export format)
    back into submission dicts.

    Returns (parsed, warnings) where:
      parsed   = { (username, entity, reporting_month): {
                      "entity": str, "reporting_month": str,
                      "activities_month": str, "status": str,
                      "initiatives": [ {...new_initiative()-shaped dicts...} ]
                  } }
      warnings = list of human-readable strings about rows/sheets that were
                 skipped or had to be guessed at.
    """
    warnings: list[str] = []
    parsed: dict = {}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {}, [f"Could not open file as an Excel workbook: {e}"]

    for ws in wb.worksheets:
        if ws.title in ("No Data",):
            continue

        hdr_row, hdrs = _find_header_row(ws)
        if not hdr_row:
            warnings.append(f"Sheet '{ws.title}': couldn't find a header row — skipped.")
            continue

        has_user_col = "user" in hdrs
        sheet_entity = _extract_entity(ws)
        fallback_username, fallback_filing = _extract_filing_and_username_from_subtitle(ws)

        if not sheet_entity:
            warnings.append(f"Sheet '{ws.title}': couldn't determine Entity — skipped.")
            continue

        col = lambda name: hdrs.get(name.lower())
        n_rows_in_sheet = 0

        for r in range(hdr_row + 1, ws.max_row + 1):
            name_col = col("initiative name")
            if not name_col:
                continue
            iname = ws.cell(r, name_col).value
            if not iname or not str(iname).strip():
                continue   # blank row

            def get(colname, default=""):
                ci = col(colname)
                if not ci:
                    return default
                v = ws.cell(r, ci).value
                return v if v is not None else default

            username = str(get("user", "")).strip() or fallback_username
            if not username:
                warnings.append(f"Sheet '{ws.title}' row {r}: no username found — skipped.")
                continue

            month_yr_label = str(get("month/yr", "")).strip()
            activities_month = parse_month_label(month_yr_label)

            filing_label = str(get("filing month", "")).strip()
            reporting_month = parse_month_label(filing_label)
            if not reporting_month:
                reporting_month = fallback_filing
            if not reporting_month and activities_month:
                # Last resort: assume the standard one-month convention
                reporting_month = next_month_of(activities_month)
                warnings.append(
                    f"Sheet '{ws.title}' row {r}: no Filing Month found — "
                    f"assumed {fmt_month(reporting_month)} (one month after activities)."
                )
            if not reporting_month:
                warnings.append(f"Sheet '{ws.title}' row {r}: couldn't determine a filing period — skipped.")
                continue

            status_key = _parse_status_label(str(get("status", "")))

            init = new_initiative()
            # Read by header name, so files exported before this column
            # existed still import — they just come back with no group.
            init["employee_group"]         = str(get(GROUP_LABEL.lower(), ""))
            init["business_component"]     = str(get("business component", ""))
            init["initiative_name"]        = str(iname).strip()
            init["initiative_description"] = str(get("initiative description", ""))
            init["tech_uncertainty"]        = str(get("tech uncertainty", ""))
            init["start_date"]              = _cell_to_date_str(get("start date", None)) or None
            init["expected_end_date"]       = _cell_to_date_str(get("expected end date", None)) or None
            init["activities"]              = str(get("activities to eliminate technical uncertainty", ""))
            team_raw                        = str(get("team members", ""))
            init["team_members"]            = [t.strip() for t in team_raw.split(",") if t.strip()]
            con_raw                         = str(get("contractors", ""))
            init["contractors"]             = [c.strip() for c in con_raw.split(",") if c.strip()]
            init["notes"]                   = str(get("notes", ""))
            init["month_yr"]                = activities_month
            init["carry_over"]              = False
            # No series column in the export, so group by name within the entity.
            init["series_id"]               = series_id_of(init)
            # A row whose month predates the report's own period is a carried-in
            # historical row, not this period's work.
            init["historical"] = bool(
                activities_month and reporting_month
                and activities_month < prev_month_of(reporting_month)
            )

            if status_key in ("approved", "archived"):
                init["initiative_status"] = "approved"
            elif status_key == "rejected":
                init["initiative_status"] = "returned"
            else:
                init["initiative_status"] = "active"

            comp_raw = get("completion date", "")
            if comp_raw and status_key in ("approved", "archived"):
                for fmt_str in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
                    try:
                        dt = datetime.strptime(str(comp_raw).strip(), fmt_str)
                        init["approved_at"] = int(dt.timestamp() * 1000)
                        break
                    except Exception:
                        continue

            key = (username, sheet_entity, reporting_month)
            if key not in parsed:
                parsed[key] = {
                    "entity":           sheet_entity,
                    "reporting_month":  reporting_month,
                    "activities_month": activities_month or prev_month_of(reporting_month),
                    "status":           status_key,
                    "initiatives":      [],
                }
            parsed[key]["initiatives"].append(init)
            n_rows_in_sheet += 1

        if n_rows_in_sheet == 0:
            warnings.append(f"Sheet '{ws.title}': no usable initiative rows found.")

    return parsed, warnings


def apply_import(parsed: dict, overwrite: bool = False) -> dict:
    """
    Commits parsed import data to storage via save_draft.
    Skips (username, entity, period) combos that already have a real
    submission unless overwrite=True.
    Returns {"imported": [...], "skipped": [...]} — lists of
    "username — Entity X — Month Year" strings.
    """
    imported, skipped = [], []
    for (username, entity, rm), data in parsed.items():
        label = f"{username} — Entity {entity} — {fmt_month(rm)}"
        existing = load_submission(username, entity, rm)
        if existing and existing.get("initiatives") and not overwrite:
            skipped.append(label)
            continue

        draft = {
            "user":             username,
            "initiatives":      data["initiatives"],
            "status":           data["status"] if data["status"] in STATUS_LABELS else "approved",
            "entity":           entity,
            "reporting_month":  rm,
            "activities_month": data["activities_month"],
        }
        # Restored reports were submitted at some point or they wouldn't exist in
        # an export, so give them a submitted_at — status is derived from it.
        if draft["status"] != "in-progress":
            draft.setdefault("submitted_at", int(time.time() * 1000))
        save_draft(username, draft)
        imported.append(label)
    return {"imported": imported, "skipped": skipped}


def get_combos() -> list[tuple[str, str]]:
    """Unique (entity, filing period) pairs that have data, oldest first."""
    return store.list_combos()


# ── Memory / history helpers ─────────────────────────────────────────────────

def get_past_ongoing_initiatives(username: str, current_reporting_month: str) -> list[dict]:
    """
    Ongoing initiatives from all of this user's earlier reports — anything whose
    expected end date is still ahead of us and that hasn't been marked resolved.
    Deduplicated by name, keeping the most recent version of each.
    """
    today  = date.today()
    seen   = set()
    result = []

    for rep in sorted(
        user_reports(username),
        key=lambda r: r.get("reporting_month", ""),
        reverse=True,
    ):
        if rep.get("reporting_month", "") >= current_reporting_month:
            continue
        for init in latest_per_series(rep):
            sid = series_id_of(init)
            if not sid or sid in seen:
                continue
            seen.add(sid)
            if init.get("resolved"):
                continue
            end = init.get("expected_end_date")
            if end:
                try:
                    if datetime.strptime(str(end), "%Y-%m-%d").date() < today:
                        continue
                except ValueError:
                    pass
            result.append(init)
    return result


def get_user_history(username: str) -> list[dict]:
    """This user's reports, newest filing period first."""
    return user_reports(username)


# ── Entity rollover ───────────────────────────────────────────────────────────

def advance_report(report: dict, target_month: str | None = None) -> tuple[bool, str, str]:
    """Archive one accepted report and create the next period's draft from it.

    This is the app's version of Power Automate Flow 4: the file is copied, the
    copy is retitled to the next period, and the original is archived. Called
    automatically the moment a report becomes approved.

    The copy KEEPS every existing row unchanged — the record of what was done in
    each earlier month — and adds one fresh row per ongoing initiative, stamped
    with the new period and with the activities field blank for the preparer.
    Resolved initiatives get no new row.

    target_month defaults to the month after the source, regardless of how far
    back the source is. Accepting a May report in September rolls it to June, not
    to October, because the periods have to be filed in sequence.

    Returns (rolled, target_month, reason_if_not_rolled).
    """
    import copy as _copy

    username     = report.get("user", "")
    entity       = report.get("entity", "")
    source_month = report.get("reporting_month", "")
    if not (username and entity and source_month):
        return False, "", "report is missing its identity"

    target_month = target_month or next_month_of(source_month)
    now          = int(time.time() * 1000)
    target_am    = prev_month_of(target_month)

    if not (report.get("initiatives") or []):
        return False, target_month, "report has no initiatives"

    # Archive the source either way — that part is not conditional on the copy.
    if report.get("status") != "archived":
        report["status"]      = "archived"
        report["archived_at"] = now
        store.save_report(report)

    existing = load_submission(username, entity, target_month)
    if existing and existing.get("initiatives"):
        return False, target_month, f"{fmt_month(target_month)} already has data"

    new_inits: list[dict] = []

    # 1. Every existing row carries over untouched, frozen as history.
    for i in report.get("initiatives") or []:
        hist = _copy.deepcopy(i)
        hist["series_id"]  = series_id_of(i)
        hist["historical"] = True
        if not hist.get("month_yr"):
            hist["month_yr"] = report.get("activities_month") or source_month
        new_inits.append(hist)

    # 2. One new row per ongoing series, for the new period.
    for series in latest_per_series(report):
        if series.get("resolved"):
            continue
        fresh = _copy.deepcopy(series)
        fresh["id"]                = f"{now}_{len(new_inits)}"
        fresh["series_id"]         = series_id_of(series)
        fresh["month_yr"]          = target_am
        fresh["historical"]        = False
        fresh["carry_over"]        = True
        fresh["activities"]        = ""
        fresh["notes"]             = ""
        fresh["pathway"]           = ""
        fresh["initiative_status"] = "active"
        fresh["approved_at"]       = None
        fresh["returned_at"]       = None
        fresh["review_comment"]    = ""
        fresh["review_history"]    = []
        fresh["completion_date"]   = None
        fresh["resolved"]          = False
        new_inits.append(fresh)

    save_draft(username, {
        "user":             username,
        "initiatives":      new_inits,
        "status":           "in-progress",
        "entity":           entity,
        "reporting_month":  target_month,
        "activities_month": target_am,
        "rolled_over_from": source_month,
    })
    return True, target_month, ""


def accept_report(report: dict) -> tuple[bool, str, str]:
    """Accept every current-period initiative, then archive and roll forward.

    Acceptance is the trigger, exactly as the status change to Complete triggered
    Flow 4 — the admin does not have to visit the Rollover tab for the normal
    monthly cycle.
    """
    now = int(time.time() * 1000)
    for i in current_rows(report):
        clear_return_flag(i, reason="accepted")
        i["initiative_status"] = "approved"
        i["approved_at"]       = now
    report["approved_at"] = now
    save_draft(report.get("user", ""), report)
    return advance_report(report)


def rollover_entity(
    source_entity: str,
    source_month: str,
    target_month: str,
    only_users: list[str] | None = None,
) -> list[str]:
    """Bulk rollover from the admin's Rollover tab.

    With acceptance now rolling reports forward on its own, this is the exception
    path: catching up after a gap, or sending a period to a month other than the
    next one.
    """
    rolled: list[str] = []
    for rep in all_reports(entity=source_entity, period=source_month):
        username = rep.get("user", "")
        if only_users is not None and username not in only_users:
            continue
        if computed_status(rep) not in ("approved", "archived"):
            continue
        did, _, _ = advance_report(rep, target_month)
        if did:
            rolled.append(username)
    return rolled


def rollable_users(entity: str, source_month: str, target_month: str) -> list[str]:
    """Users who would actually be rolled by the given move."""
    out = []
    for rep in all_reports(entity=entity, period=source_month):
        if computed_status(rep) not in ("approved", "archived"):
            continue
        if not (rep.get("initiatives") or []):
            continue
        username = rep.get("user", "")
        existing = load_submission(username, entity, target_month)
        if existing and existing.get("initiatives"):
            continue
        out.append(username)
    return sorted(out)


# ── Initiative helpers ────────────────────────────────────────────────────────

def new_initiative() -> dict:
    import random as _rnd
    return {
        "id": f"{int(datetime.now().timestamp()*1000)}_{_rnd.randint(10000,99999)}",
        "employee_group":        "",
        "business_component":    "",
        "initiative_name":       "",
        "initiative_description":"",
        "tech_uncertainty":      "",
        "start_date":            None,
        "expected_end_date":     None,
        "activities":            "",
        "team_members":          [],
        # contractors: people who are NOT in the employee directory — outside
        # firms, and anyone missing from the HR export. Kept in their own field
        # rather than mixed into team_members so the two stay distinguishable in
        # the export no matter what happens to the directory afterwards.
        "contractors":           [],
        "notes":                 "",
        "carry_over":            False,
        # Per-initiative status (set by admin actions)
        "initiative_status":     "active",   # active | approved | returned
        "approved_at":           None,       # ms timestamp when admin approved
        "returned_at":           None,       # ms timestamp when admin returned
        # review_comment: the outstanding reviewer note, shown only while the
        # initiative is returned. Cleared into review_history once revised.
        "review_comment":        "",
        "review_history":        [],
        # series_id: stable across months. Each month's row shares it.
        "series_id":             f"s{int(datetime.now().timestamp()*1000)}_{_rnd.randint(10000,99999)}",
        # historical: True for a row carried in from an earlier month (read-only).
        "historical":            False,
        # month_yr: the reporting period this initiative row belongs to.
        # Set when first saved; preserved through admin rollovers so each row
        # always shows the correct original period in the export.
        "month_yr":              "",
        # pathway: which flow the user chose this month for this carry-over initiative.
        # "continuing" | "resolved" | "new_direction" | "" (not yet chosen)
        "pathway":               "",
        # resolved: the TECHNICAL UNCERTAINTY has been resolved. Set via the
        # Resolved pathway. Resolved initiatives get no new row in later periods
        # and drop out of carry-over suggestions.
        #
        # Distinct from closing the BUSINESS COMPONENT, which is tracked per
        # (entity, business component) in config — see close_business_component().
        # An uncertainty can resolve while the component keeps running.
        "resolved":              False,
        # completion_date: user-set date when resolved (YYYY-MM-DD).
        # Populates the Completion Date column in the export.
        "completion_date":       None,
    }

def split_legacy_members(init: dict) -> dict:
    """Backfill `contractors` on initiatives saved before the split existed.

    Older rows kept contractors inside team_members. Splitting on the directory
    is only safe to do ONCE, at which point the result is stored: if it were
    recomputed on every read, replacing or removing the directory would silently
    reclassify saved employees as contractors and change historical exports.
    """
    if "contractors" in init:
        return init
    members = init.get("team_members") or []
    if directory.is_loaded():
        init["team_members"] = [m for m in members if directory.known(m)]
        init["contractors"]  = [m for m in members if not directory.known(m)]
    else:
        # No directory to judge against — leave everything as team members
        # rather than guess.
        init["contractors"] = []
    return init


def employees_of(init: dict) -> list[str]:
    split_legacy_members(init)
    return list(init.get("team_members") or [])


def contractors_of(init: dict) -> list[str]:
    split_legacy_members(init)
    return list(init.get("contractors") or [])


def all_people_on(init: dict) -> list[str]:
    """Everyone on an initiative, for display where the distinction doesn't matter."""
    return employees_of(init) + contractors_of(init)


def carryover_initiative(src: dict) -> dict:
    """Manual carry-over from the dashboard. Continues the same series, so the
    export groups it with the earlier months of the same initiative."""
    init = new_initiative()
    for f in CARRYOVER_FIELDS:
        init[f] = src.get(f, init[f])
    init["carry_over"] = True
    init["series_id"]  = series_id_of(src)
    init["historical"] = False
    return init

def empty_draft() -> dict:
    return {
        "initiatives":      [],
        "status":           "in-progress",
        "entity":           "",
        "reporting_month":  "",   # filing month  — storage key & Group# in filename
        "activities_month": "",   # activities month — what period the R&D work took place in
    }


# ── Excel export ──────────────────────────────────────────────────────────────
# Colors from the original template (extracted via openpyxl):
#   GREEN  = 9BBB59  (theme accent3 — header row)
#   BEIGE  = EEECE1  (theme lt2   — input cells)
#   YELLOW = DED900  (rgb         — Team Members / Selection cells)
#   Font   = Arial Narrow, bold white 12pt on headers

GREEN  = "9BBB59"
BEIGE  = "EEECE1"
YELLOW = "DED900"
WHITE  = "FFFFFF"
DARK   = "1F1F1F"
GRAY   = "666666"

def _thin_border():
    s = Side(style="thin", color="A0A0A0")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font(sz=12):
    return Font(name="Arial Narrow", bold=True, color=WHITE, size=sz)

def _data_font():
    return Font(name="Arial Narrow", color=DARK, size=11)

# Column definition for a single-combo sheet (User prepended for consolidated)
_COL_DEF = [
    # Employee Group leads, matching the order the app asks for things. Every
    # other width matches the original template exactly (measured via openpyxl
    # from the source file).
    (GROUP_LABEL,                                              "employee_group",     32.0,   False),
    ("Month/Yr",                                               "_month",             14.15,  False),
    ("Filing Month",                                            "_filing",            13.0,   False),
    ("Business Component",                                     "business_component", 35.0,   False),
    ("Component Status",                                       "_bc_status",         24.0,   False),
    ("Initiative Name",                                        "initiative_name",    19.26,  False),
    ("Initiative Description",                                 "initiative_description", 38.15, False),
    ("Tech Uncertainty",                                       "tech_uncertainty",   64.41,  False),
    ("Start Date",                                             "start_date",         15.68,  False),
    ("Expected End Date",                                      "expected_end_date",  15.68,  False),
    ("Activities to Eliminate Technical Uncertainty",          "activities",         60.26,  False),
    ("Team Members",                                           "team_members",       49.0,   True),
    ("Contractors",                                            "contractors",        30.0,   True),
    ("Notes",                                                  "notes",              52.57,  False),
    ("Completion Date",                                        "_completion",        22.0,   False),
    ("Status",                                                 "_status",            22.0,   False),
]

def _write_sheet(ws, rows_data: list[dict], subtitle: str):
    """
    Matches the exact structure of the original template:
      Row 1 : Title (A1) + Legend header cells (C-E or D-F for consolidated)
      Row 2 : Legend swatches (Selection=yellow, subtitle text)
      Row 3 : Autofill label
      Row 4 : Empty spacer
      Row 5 : Column headers (green, Arial Narrow 12pt bold white)
      Row 6+: Data rows (auto height, beige fill / yellow for Team Members)
    """
    border      = _thin_border()
    fill_green  = PatternFill("solid", fgColor=GREEN)
    fill_beige  = PatternFill("solid", fgColor=BEIGE)
    fill_yellow = PatternFill("solid", fgColor=YELLOW)
    fill_red    = PatternFill("solid", fgColor="FF0000")

    has_user_col = any("user" in r for r in rows_data) if rows_data else False
    cols = ([("User", "user", 18, False)] if has_user_col else []) + list(_COL_DEF)

    # Legend columns offset right by 1 for consolidated sheet (extra User col)
    lc = 3 if not has_user_col else 4   # start column for legend

    # ── Row 1: Title + Legend header ─────────────────────────────────────────
    ws.row_dimensions[1].height = 18.3
    c = ws.cell(1, 1, "Monthly R&D Tracking Template")
    c.font = Font(name="Arial Narrow", bold=True, size=14, color=DARK)
    c = ws.cell(1, lc, "Legend:")
    c.font = Font(name="Arial Narrow", bold=True, size=12, color=DARK)
    c.alignment = Alignment(horizontal="right", wrap_text=True)
    c = ws.cell(1, lc + 1, "Input")
    c.font = Font(name="Arial Narrow", size=12, color=DARK)
    c.fill = fill_beige
    c.alignment = Alignment(wrap_text=True)
    c = ws.cell(1, lc + 2, "Missing Info")
    c.font = Font(name="Arial Narrow", size=11, color=DARK)
    c.fill = fill_red

    # ── Row 2: Selection swatch + subtitle ───────────────────────────────────
    ws.row_dimensions[2].height = 15.3
    c = ws.cell(2, lc + 1, "Selection")
    c.font = Font(name="Arial Narrow", size=12, color=DARK)
    c.fill = fill_yellow
    c.alignment = Alignment(wrap_text=True)
    c = ws.cell(2, lc + 2, subtitle)
    c.font = Font(name="Arial Narrow", size=10, color=GRAY, italic=True)

    # ── Row 3: Autofill label ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 15.3
    c = ws.cell(3, lc + 1, "Autofill")
    c.font = Font(name="Arial Narrow", bold=True, size=12, color=DARK)
    c.alignment = Alignment(wrap_text=True)

    # ── Row 4: Empty spacer ───────────────────────────────────────────────────
    ws.row_dimensions[4].height = 15

    # ── Row 5: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[5].height = 30
    for ci, (hdr, _, width, _) in enumerate(cols, 1):
        cell = ws.cell(5, ci, hdr)
        cell.fill      = fill_green
        cell.font      = _hdr_font(12)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── Row 6+: Data rows ─────────────────────────────────────────────────────
    # No explicit height — Excel auto-fits to content (matches original template behavior)
    for rn, row in enumerate(rows_data, 6):
        for ci, (_, key, _, is_team) in enumerate(cols, 1):
            val = row.get(key, "")
            cell = ws.cell(rn, ci, val)
            cell.font      = _data_font()
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill      = fill_yellow if is_team else fill_beige

    # Freeze header rows 1-5 (same as original)
    ws.freeze_panes = "A6"


def _bc_status_cell(entity: str, bc: str) -> str:
    """Whether this row's business component is open or closed.

    The closure DATE is written out rather than a bare "Closed", because
    closure is current state while each row belongs to a past month. A March row
    under a component closed in June is not a March closure, and a bare "Closed"
    would read as one. With the date present, it can be compared against the
    row's own Month/Yr two columns to the left.
    """
    rec = bc_is_closed(entity, bc)
    if not (bc or "").strip():
        return ""
    if not rec:
        return "Open"
    when = ts_to_et(rec.get("closed_at"), "%Y-%m-%d")
    return f"Closed {when}" if when else "Closed"


def _sub_to_rows(username: str, sub: dict, include_user: bool) -> list[dict]:
    """Convert a submission dict into a list of row dicts for _write_sheet."""
    rm     = sub.get("reporting_month", "")
    entity = sub.get("entity", "")
    status = STATUS_LABELS.get(computed_status(sub), sub.get("status", ""))
    rows   = []
    for init in sub.get("initiatives") or []:
        row = {
            # Use initiative's own month_yr, then activities_month, then filing month
            "_month":  fmt_month(init.get("month_yr") or sub.get("activities_month") or rm),
            "_filing": fmt_month(rm),
            "_status": status,
            # Completion date: the user's own date if they resolved the
            # initiative, otherwise when it was accepted. Both formatted in
            # Eastern time via the same helper, rather than one going through a
            # fixed -5 offset and the other through the server's local timezone.
            # Completion Date means the technical uncertainty was RESOLVED.
            # It used to fall back to the approval date, which stamped a
            # completion date on every accepted initiative including ones still
            # in progress — an examiner reading the file literally would have
            # seen ongoing work reported as finished.
            "_completion": (init.get("completion_date") or "") if init.get("resolved") else "",
            "employee_group":        init.get("employee_group",        ""),
            "business_component":    init.get("business_component",    ""),
            "_bc_status":            _bc_status_cell(entity, init.get("business_component", "")),
            "initiative_name":       init.get("initiative_name",        ""),
            "initiative_description":init.get("initiative_description", ""),
            "tech_uncertainty":      init.get("tech_uncertainty",       ""),
            "start_date":            str(init.get("start_date")        or ""),
            "expected_end_date":     str(init.get("expected_end_date") or ""),
            "activities":            init.get("activities",             ""),
            "team_members":          ", ".join(employees_of(init)),
            "contractors":           ", ".join(contractors_of(init)),
            "notes":                 init.get("notes",                  ""),
        }
        if include_user:
            row["user"] = username
        rows.append(row)
    return rows


def build_excel_individual(username: str, sub: dict) -> bytes:
    """Single-user, single-month export — matches original template layout."""
    wb = Workbook()
    ws = wb.active
    am     = sub.get("activities_month") or sub.get("reporting_month", "")
    fm     = sub.get("reporting_month", "")
    entity = sub.get("entity", "")
    ws.title = f"Entity {entity}" if entity else (fmt_month_tab(am) or "Report")
    subtitle = (
        f"Submitted by: {username}  |  Entity: {entity}  |  "
        f"Filing: {fmt_month(fm)}  |  Reporting Period: {fmt_month(am)}"
    )
    rows, _ = collect_rows([{**sub, "user": username}], include_user=False)
    rows.sort(key=lambda r: (r.get("_month", ""), r.get("initiative_name", "")))
    _write_sheet(ws, rows, subtitle)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_consolidated(
    reports: list[dict],
    group_by: str = "entity",   # "entity" → one tab per entity | "user" → one tab per person
    status_desc: str = "",
    dedupe: bool = True,
) -> bytes:
    """
    Consolidated workbook from a list of reports. Callers filter the list; this
    function only lays it out.

    group_by="entity"  one sheet per entity, rows sorted by period then user
    group_by="user"    one sheet per person, rows sorted by entity then period

    dedupe=True (deliverable): one row per initiative per month, duplicates from
    carried-forward copies collapsed. This is what you hand to a reviewer.

    dedupe=False (backup): every row from every period file, duplicates included.
    Each row carries its own Filing Month, so the restore can rebuild the separate
    period files exactly. Deduplicating a backup would silently discard the
    intermediate periods, because the newest report's copy would win.
    """
    wb = Workbook()
    wb.remove(wb.active)
    status_desc = status_desc or "All statuses"

    if group_by == "user":
        for username in sorted({r.get("user", "") for r in reports if r.get("user")}):
            mine = [r for r in reports if r.get("user") == username]
            rows, _ = collect_rows(mine, include_user=False, dedupe=dedupe)
            rows.sort(key=lambda r: (r.get("_month", ""), r.get("initiative_name", "")))
            if not rows:
                continue
            ws = wb.create_sheet(title=(username or "Unknown")[:31])
            _write_sheet(ws, rows, f"Team member: {username}  |  {status_desc}")

    else:
        for entity in sorted({r.get("entity", "") for r in reports if r.get("entity")}):
            theirs = [r for r in reports if r.get("entity") == entity]
            rows, _ = collect_rows(theirs, include_user=True, dedupe=dedupe)
            rows.sort(key=lambda r: (r.get("_month", ""), r.get("user", ""), r.get("initiative_name", "")))
            if not rows:
                continue
            ws = wb.create_sheet(title=f"Entity {entity}"[:31])
            _write_sheet(ws, rows, f"Entity: {entity}  |  {status_desc}")

    if not wb.sheetnames:
        ws = wb.create_sheet("No Data")
        ws.cell(1, 1, "No reports match the selected filters.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def collect_rows(
    reports: list[dict], include_user: bool, dedupe: bool = True
) -> tuple[list[dict], int]:
    """Flatten reports into export rows, dropping duplicates.

    Because rolling forward copies earlier rows into the new period's report, the
    same (initiative, month) appears in every later report as well as its own.
    Selecting several periods would otherwise list those rows once per report.

    Rows are keyed on (user, entity, series, month). Where the same row appears
    more than once, the version from the LATEST report wins, since that is the
    most recently reviewed copy. Returns the rows plus how many duplicates were
    dropped, so the UI can say so rather than quietly changing the count.
    """
    if not dedupe:
        out = []
        for rep in sorted(reports, key=lambda r: r.get("reporting_month", "")):
            out.extend(_sub_to_rows(rep.get("user", ""), rep, include_user=include_user))
        return out, 0

    seen: dict[tuple, dict] = {}
    dropped = 0
    # Ascending, so later reports overwrite earlier copies of the same row.
    for rep in sorted(reports, key=lambda r: r.get("reporting_month", "")):
        username = rep.get("user", "")
        for row, init in zip(
            _sub_to_rows(username, rep, include_user=include_user),
            rep.get("initiatives") or [],
        ):
            key = (
                username,
                rep.get("entity", ""),
                series_id_of(init),
                init.get("month_yr") or rep.get("activities_month") or "",
            )
            if key in seen:
                dropped += 1
            seen[key] = row
    return list(seen.values()), dropped


def dedupe_stats(reports: list[dict]) -> tuple[int, int]:
    """(unique rows, duplicates dropped) for the export summary."""
    rows, dropped = collect_rows(reports, include_user=True)
    return len(rows), dropped


def status_desc_for(status_keys: list[str]) -> str:
    """Human-readable status list for an export subtitle."""
    if not status_keys:
        return "None selected"
    return ", ".join(
        STATUS_LABELS.get(s, s).split(" ", 1)[-1] for s in status_keys
    )


# ── CSS ───────────────────────────────────────────────────────────────────────

# ── Employee group and team selection ─────────────────────────────────────────
# Every employee in the uploaded directory belongs to exactly one group (the
# client's Cost Center column), so there is no multi-assignment ambiguity. For
# the current client that is 152 groups across 2,132 people, median 7 each —
# the difference between a usable dropdown and a useless one.
#
# The group is chosen FIRST, stored on the initiative, and the team list is then
# strictly that group's people. There is no way to reach across groups: work
# spanning two groups is two initiatives. Write-ins are the one exception, since
# a contractor belongs to no group by definition.
#
# With no directory uploaded, group selection falls back to free text and team
# selection to the built-in EMPLOYEES list, so a fresh install still works.

def group_select(state_key: str, current: str = "", label: str | None = None) -> str:
    """Choose an employee group. Returns the chosen group, "" if none."""
    label = label or GROUP_LABEL
    if not directory.is_loaded():
        st.caption(
            "No employee directory loaded, so this is free text and team "
            "selection falls back to the built-in list. An admin can upload the "
            "client's directory in **Admin → Settings**."
        )
        return st.text_input(label, value=current or "", key=f"{state_key}__free")

    opts = directory.groups()
    n_emp, n_grp = directory.counts()
    placeholder = f"— Choose a {label.lower()} —"
    choices = [placeholder] + opts
    if current and current not in opts:
        # A group that has since been renamed or dropped from the directory.
        # Keep it selectable rather than silently reassigning the initiative.
        choices = [placeholder, current] + opts
    idx = choices.index(current) if current in choices else 0

    picked = st.selectbox(
        label, choices, index=idx, key=f"{state_key}__grp",
        help=f"{n_emp:,} employees across {n_grp} groups. Type to search — the "
             "code and the name both match.",
    )
    return "" if picked == placeholder else picked


def _prune_to_group(members, group: str):
    """Drop anyone who isn't in `group`. Write-ins are kept — a contractor
    isn't in the directory at all, so they belong to whichever initiative
    someone added them to."""
    return [m for m in (members or [])
            if not directory.known(m) or directory.group_of(m) == group]


def render_team_picker(state_key: str, group: str, current=None,
                       label: str = "Team Members", current_contractors=None):
    """Select team members from within one group, plus any contractors.

    Returns (employees, contractors) and also writes them to
    st.session_state[state_key] and st.session_state[state_key + "__con"], so
    callers can read them whichever way suits them.

    The two are kept apart from the moment of entry rather than separated later
    on the way to the export. Deriving the split at export time would make it
    depend on whichever directory happened to be loaded that day.
    """
    con_key = f"{state_key}__con"
    if state_key not in st.session_state:
        st.session_state[state_key] = list(current or [])
    if con_key not in st.session_state:
        st.session_state[con_key] = list(current_contractors or [])
    selected  = list(st.session_state[state_key])
    write_ins = list(st.session_state[con_key])

    # ── No directory — built-in list ─────────────────────────────────────────
    if not directory.is_loaded():
        ms_key = f"{state_key}__builtin"
        if ms_key not in st.session_state:
            st.session_state[ms_key] = [m for m in selected if m in EMPLOYEES]
        picked = st.multiselect(
            label, EMPLOYEES, key=ms_key, placeholder="Select team members...",
        )
        # Write-ins and directory names aren't in EMPLOYEES; keep them rather
        # than silently dropping them.
        extras = [m for m in selected if m not in EMPLOYEES]
        result = list(picked) + extras
        st.session_state[state_key] = result
        return result, write_ins

    if not group:
        st.info(f"Choose a {GROUP_LABEL.lower()} first — team members come from it.")
        return selected, write_ins

    people = directory.names_in(group)
    ms_key = f"{state_key}__ms"
    if ms_key not in st.session_state:
        st.session_state[ms_key] = [p for p in selected if p in people]
    picked = st.multiselect(
        f"{label} — {len(people)} in {group}",
        people,
        key=ms_key,
        placeholder="Select team members...",
    )

    # ── Contractors and anyone missing from the file ───────────────────────
    with st.expander(f"Someone not in the directory ({len(write_ins)} added)"
                     if write_ins else "Someone not in the directory (contractor, new hire)"):
        st.caption(
            "The directory is employees only, so contractors have to be added "
            "here. Give the company too and it is stored as *Name (Company)*, "
            "which keeps the entries consistent across preparers."
        )
        wc1, wc2 = st.columns(2)
        with wc1:
            wi_name = st.text_input("Name", key=f"{state_key}__wi_name")
        with wc2:
            wi_co = st.text_input(
                "Company (contractors only)", key=f"{state_key}__wi_co",
                placeholder="Leave blank for an employee",
            )
        if st.button("Add to team", key=f"{state_key}__wi_add"):
            nm, co = (wi_name or "").strip(), (wi_co or "").strip()
            if not nm:
                st.warning("Enter a name first.")
            else:
                entry = f"{nm} ({co})" if co else nm
                if entry in write_ins:
                    st.info(f"{entry} is already on this initiative.")
                else:
                    st.session_state[con_key] = write_ins + [entry]
                    st.session_state.pop(f"{state_key}__wi_name", None)
                    st.session_state.pop(f"{state_key}__wi_co", None)
                    st.rerun()

        for person in write_ins:
            wc3, wc4 = st.columns([8, 1])
            with wc3:
                st.markdown(
                    f"{person}  \n<small style='color:#64748b;'>"
                    "recorded in the Contractors column</small>",
                    unsafe_allow_html=True,
                )
            with wc4:
                if st.button("✕", key=f"{state_key}__rm_{abs(hash(person)) % 10**8}",
                             help=f"Remove {person}"):
                    st.session_state[con_key] = [p for p in write_ins if p != person]
                    st.rerun()

    st.session_state[state_key] = list(picked)
    st.session_state[con_key]   = write_ins
    return list(picked), write_ins


def clear_team_picker_state(state_key: str):
    """Drop a picker's internal widget state so the next initiative doesn't
    inherit the last one's selection."""
    for k in [k for k in list(st.session_state) if k.startswith(f"{state_key}__")]:
        st.session_state.pop(k, None)
    st.session_state.pop(f"{state_key}__con", None)


def inject_css():
    st.markdown("""
    <style>
    #MainMenu, footer, header { visibility: hidden; }
    .stApp { background: #f1f5f9; }
    .badge-approved    { background:#f0fdf4; color:#166534; border:1.5px solid #86efac; }
    .badge-submitted   { background:#eff6ff; color:#1e40af; border:1.5px solid #93c5fd; }
    .badge-in-progress { background:#fef9ec; color:#92600a; border:1.5px solid #fcd34d; }
    .badge-rejected    { background:#fff7ed; color:#9a3412; border:1.5px solid #fdba74; }
    .badge-archived    { background:#f8fafc; color:#475569; border:1.5px solid #cbd5e1; }
    .badge-not-started { background:#f8fafc; color:#64748b; border:1.5px solid #cbd5e1; }
    .rd-badge { display:inline-block; padding:3px 12px; border-radius:20px; font-size:12px; font-weight:700; }
    .progress-bar-outer { background:#e2e8f0; border-radius:4px; height:8px; overflow:hidden; margin-bottom:6px; }
    .progress-bar-inner { background:#c86a2a; height:100%; border-radius:4px; transition:width 0.3s; }
    .wizard-question { font-size:22px; font-weight:700; color:#1a3c5e; margin-bottom:6px; }
    .wizard-hint     { font-size:14px; color:#64748b; margin-bottom:20px; }
    .step-label      { font-size:11px; font-weight:700; color:#c86a2a; text-transform:uppercase; letter-spacing:.8px; }
    .carryover-banner { background:#fffbeb; border:1.5px solid #fcd34d; border-radius:10px; padding:10px 16px; margin-bottom:14px; font-size:13px; color:#92600a; }
    .prefilled-banner { background:#f0f9ff; border:1.5px solid #93c5fd; border-radius:10px; padding:10px 16px; margin-bottom:14px; font-size:13px; color:#1e40af; }
    .delete-confirm   { background:#fff1f2; border:1.5px solid #fda4af; border-radius:10px; padding:12px 16px; margin-top:8px; }
    </style>
    """, unsafe_allow_html=True)

def badge_html(status: str) -> str:
    cls_map = {
        "approved":    "badge-approved",
        "submitted":   "badge-submitted",
        "in-progress": "badge-in-progress",
        "rejected": "badge-rejected",
    }
    cls   = cls_map.get(status, "badge-not-started")
    label = STATUS_LABELS.get(status, status)
    return f'<span class="rd-badge {cls}">{label}</span>'

def best_draft_for_user(username: str) -> dict:
    """
    Picks the report to open on sign-in, in priority order:
      1. The earliest in-progress report with real content — usually a rollover
         waiting to be reviewed and submitted.
      2. The current filing month, if it already exists.
      3. The most recent report with any content.
      4. A fresh empty draft for the current month.
    """
    reports = user_reports(username)

    in_progress = sorted(
        [r for r in reports if r.get("status") == "in-progress"],
        key=lambda r: r.get("reporting_month", ""),
    )
    if in_progress:
        return in_progress[0]

    cur = cur_month()
    for r in reports:
        if r.get("reporting_month") == cur:
            return r

    if reports:
        return max(reports, key=lambda r: r.get("reporting_month", ""))

    draft = empty_draft()
    draft["reporting_month"]  = cur
    draft["activities_month"] = prev_month_of(cur)
    return draft


def init_session():
    defaults = {
        "screen":      "login",
        "user":        None,
        "is_admin":    False,
        "draft":       empty_draft(),
        "wiz_init":    None,
        "wiz_step":    0,
        "wiz_mode":    "new",
        "confirm_del": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


# ── Login ─────────────────────────────────────────────────────────────────────

def screen_login():
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align:center; margin-bottom:28px;">
            <div style="font-size:48px; margin-bottom:10px;">🔬</div>
            <h1 style="color:#1a3c5e; font-size:28px; margin:0 0 6px;">R&D Activity Tracker</h1>
            <p style="color:#64748b; margin:0;">Monthly Reporting Portal</p>
        </div>
        """, unsafe_allow_html=True)
        options = ["— Select your name —", "⚙ Admin (Oversight Lead)"] + EMPLOYEES
        sel = st.selectbox("Sign in as", options, label_visibility="collapsed")
        if st.button("Sign In →", width='stretch', type="primary"):
            if sel.startswith("— "):
                st.warning("Please select your name.")
            else:
                name     = sel.replace("⚙ ", "").replace(" (Oversight Lead)", "").strip()
                is_admin = "Admin" in sel
                st.session_state.user     = name
                st.session_state.is_admin = is_admin
                # Auto-load the most actionable period (in-progress rollover first,
                # then current month, then most recent) so the user lands on the
                # right report without having to change Report Setup manually.
                if "Admin" not in sel:
                    best = best_draft_for_user(name)
                    st.session_state.draft = best
                    # Clear Report Setup widget keys so Streamlit renders
                    # the selectboxes from the new draft's values rather
                    # than whatever was cached from a prior session/period.
                    for k in ["su_entity", "su_filing"]:
                        st.session_state.pop(k, None)
                    # Clear the dynamic Reporting Period key too
                    rm = best.get("reporting_month", "")
                    st.session_state.pop(f"su_act_{rm}", None)
                else:
                    st.session_state.draft = empty_draft()
                st.session_state.screen   = "admin" if is_admin else "dashboard"
                st.rerun()
        st.markdown(
            '<p style="text-align:center;font-size:12px;color:#94a3b8;margin-top:16px;">'
            'Your entries are saved automatically as you go.</p>',
            unsafe_allow_html=True,
        )


# ── Dashboard ─────────────────────────────────────────────────────────────────

# ── Reminder banners ──────────────────────────────────────────────────────────
# Streamlit has no background scheduler — these checks run only when someone
# opens the app, and simply surface a visible nudge based on today's date.

def outstanding_review_notes(report: dict) -> list[str]:
    """Reviewer notes on initiatives still awaiting revision, one per line."""
    notes = []
    for init in outstanding_returns(report):
        comment = (init.get("review_comment") or "").strip()
        name = init.get("initiative_name", "Initiative")
        notes.append(f"{name}: {comment}" if comment else name)
    return notes


def render_user_reminder(user: str, current_filing_month: str = ""):
    """
    Alerts in priority order:
      1. Reports with initiatives sent back for revision.
      2. Unsubmitted reports from other filing periods.
      3. Monthly filing nudge after the 5th.
    """
    reports = user_reports(user)

    # ── Reports needing revision ─────────────────────────────────────────────
    for rep in reports:
        if computed_status(rep) != "rejected":
            continue
        month = rep.get("reporting_month", "")
        if month == current_filing_month:
            continue   # shown inline in the status strip instead
        n = len(outstanding_returns(rep))
        st.warning(
            f"🔴 **{fmt_month(month)}** (Entity {rep.get('entity','—')}) — "
            f"{n} initiative{'s' if n != 1 else ''} need{'' if n != 1 else 's'} revision. "
            f"Switch to that period in Report Setup to update {'them' if n != 1 else 'it'}."
        )

    # ── Unsubmitted drafts elsewhere ─────────────────────────────────────────
    pending = [
        rep for rep in reports
        if computed_status(rep) == "in-progress"
        and rep.get("reporting_month") != current_filing_month
    ]
    if pending:
        parts = [
            f"**{fmt_month(r.get('reporting_month',''))}**"
            + (" *(rolled forward by the Oversight Lead)*" if r.get("rolled_over_from") else "")
            for r in pending
        ]
        st.warning(
            f"⚠ You have an unsubmitted report for {', '.join(parts)}. "
            "Open Report Setup, select that filing month, review your initiatives, "
            "then submit for review."
        )

    # ── Monthly nudge ────────────────────────────────────────────────────────
    if date.today().day >= 5:
        cur = cur_month()
        if not any(r.get("reporting_month") == cur for r in reports):
            st.info(
                f"📅 It's past the 5th and your **{fmt_month(cur)}** filing "
                "hasn't been started. Set up your report below to begin."
            )


def render_admin_reminder():
    """Surfaces a count of reports still waiting on review."""
    pending = [
        (r.get("user", ""), r.get("reporting_month", ""))
        for r in all_reports(statuses=["submitted"])
    ]
    if pending:
        names = ", ".join(f"{u} ({fmt_month(m)})" for u, m in pending[:5])
        extra = f" +{len(pending)-5} more" if len(pending) > 5 else ""
        st.warning(f"🔔 {len(pending)} report(s) pending your review: {names}{extra}")


def pathways_pending(report: dict) -> list[dict]:
    """Carried-forward initiatives in this period with no pathway chosen yet.

    A rollover stamps carry_over=True and blanks the pathway on each new row, so
    "needs an answer this period" is exactly those two conditions. Once the
    preparer picks a pathway and saves, the row drops out of this list and stays
    out until the next period blanks it again.
    """
    return [
        i for i in current_rows(report)
        if i.get("carry_over") and not (i.get("pathway") or "").strip()
    ]


def next_pathway_prompt(report: dict) -> dict | None:
    """The next initiative to ask about, skipping any deferred this session."""
    skipped = st.session_state.get("pathway_skipped") or set()
    for init in pathways_pending(report):
        if init.get("id") not in skipped:
            return init
    return None


def screen_dashboard():
    user      = st.session_state.user
    draft     = st.session_state.draft
    # Status is derived from the initiatives, not stored independently.
    # "rejected" is deliberately NOT locked — the user has to be able to
    # fix what was flagged. There is no acknowledgement step: it used to live in
    # st.session_state, which meant signing out or an app restart forced the user
    # to acknowledge the same feedback again.
    status      = computed_status(draft)
    locked      = status in LOCKED_STATUSES
    is_archived = status == "archived"
    submitted   = locked   # retained name, used widely below

    # Header
    c1, c2 = st.columns([5, 1])
    with c1:
        fm  = draft.get("reporting_month", "")
        am  = draft.get("activities_month", "")
        st.markdown("## 🔬 R&D Tracker")
        if fm and am:
            st.markdown(
                f"**Filing:** {fmt_month(fm)} &nbsp;|&nbsp; "
                f"**Reporting Period:** {fmt_month(am)}",
                unsafe_allow_html=True,
            )
        elif fm:
            st.markdown(f"**Filing Month:** {fmt_month(fm)}", unsafe_allow_html=True)
        else:
            st.markdown(
                "<span style='color:#c86a2a;'>⚠ Complete Report Setup below</span>",
                unsafe_allow_html=True,
            )
        st.caption(f"Signed in as **{user}**")
    with c2:
        st.write("")
        if st.button("Sign Out"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    st.divider()
    render_user_reminder(user, current_filing_month=draft.get("reporting_month", ""))

    # ── Rollover notification ─────────────────────────────────────────────────
    # Show a clear banner when the current draft was created by admin rollover
    # and hasn't been submitted yet — so the user knows exactly what happened.
    rolled_from = draft.get("rolled_over_from")
    if rolled_from and draft.get("status") == "in-progress":
        n_new  = len(current_rows(draft))
        n_past = len(historical_rows(draft))
        st.info(
            f"📋 **Your {fmt_month(rolled_from)} report has been rolled forward to "
            f"{fmt_month(draft.get('reporting_month',''))}.**  \n"
            f"{n_new} ongoing initiative{'s' if n_new != 1 else ''} "
            f"{'have' if n_new != 1 else 'has'} a fresh row for this month with the "
            "activities field blank — describe what you did this month, then submit. "
            + (
                f"The {n_past} row{'s' if n_past != 1 else ''} from earlier months "
                "are kept in the report as a record and can't be edited."
                if n_past else ""
            ),
            icon="📋",
        )

    # ── Report Setup ─────────────────────────────────────────────────────────
    setup_complete = bool(draft.get("entity") and draft.get("reporting_month"))
    filing_lbl     = fmt_month(draft.get("reporting_month",""))
    act_lbl        = fmt_month(draft.get("activities_month",""))
    expander_label = (
        f"⚙ Report Setup  —  Entity {draft['entity']}  ·  Filing: {filing_lbl}  ·  Reporting Period: {act_lbl}"
        if setup_complete else "⚙ Report Setup  ⚠ Please complete setup first"
    )
    with st.expander(expander_label, expanded=not setup_complete):
        st.caption(
            "**Filing Month** is the month you are submitting this report in "
            "— it determines the Group number in the filename (e.g. filing in June → Group 6). "
            "**Reporting Period** is the month the R&D activities actually took place "
"— this is the period that shows in the export (typically the previous month)."
        )
        su1, su2, su3 = st.columns(3)
        mlist = available_months()
        # If the current draft is for a period outside the normal window
        # (e.g. a rollover created 2+ months ahead), include it so the
        # dropdown doesn't fall back to the wrong month.
        cur_draft_rm = draft.get("reporting_month")
        if cur_draft_rm and cur_draft_rm not in mlist:
            mlist = sorted(set(mlist) | {cur_draft_rm}, reverse=True)

        with su1:
            entity_options = get_user_entities(user) + ["+ Add new entity..."]
            eidx = entity_options.index(draft["entity"]) if draft.get("entity") in entity_options else 0
            entity_pick = st.selectbox("Entity", entity_options, index=eidx, key="su_entity")

            if entity_pick == "+ Add new entity...":
                new_e_col1, new_e_col2 = st.columns([3, 1])
                with new_e_col1:
                    new_entity_input = st.text_input(
                        "New entity number", key="su_new_entity",
                        placeholder="e.g. 111", label_visibility="collapsed",
                    )
                with new_e_col2:
                    if st.button("Add", key="su_add_entity"):
                        if new_entity_input.strip():
                            added = add_custom_entity(new_entity_input.strip())
                            st.session_state.su_entity = new_entity_input.strip()
                            if added:
                                st.success(f"Entity {new_entity_input.strip()} added — it will stay in the dropdown.")
                            st.rerun()
                # Keep previous entity selected until a new one is actually added
                chosen_entity = draft.get("entity") or (all_entities()[0] if all_entities() else "")
            else:
                chosen_entity = entity_pick

        with su2:
            # Filing Month — defaults to CURRENT month (you are filing now)
            def_filing = mlist[0]
            cur_filing  = draft.get("reporting_month") or def_filing
            filing_idx  = mlist.index(cur_filing) if cur_filing in mlist else 0
            chosen_filing = st.selectbox(
                "Filing Month",
                mlist,
                index=filing_idx,
                format_func=fmt_month,
                key="su_filing",
                help="The month you are submitting this report in. Sets the Group number in the filename.",
            )

        with su3:
            filing_changed_now = chosen_filing != draft.get("reporting_month", "")
            if filing_changed_now:
                # Auto-derive from new filing month
                cur_act = prev_month_of(chosen_filing)
            else:
                # Filing month unchanged — respect whatever the user has chosen
                cur_act = draft.get("activities_month") or prev_month_of(chosen_filing)
            act_idx = mlist.index(cur_act) if cur_act in mlist else (1 if len(mlist) > 1 else 0)
            # Key includes the filing month — when filing changes, Streamlit treats
            # this as a completely new widget and renders from index, not cached state.
            chosen_act = st.selectbox(
                "Reporting Period",
                mlist,
                index=act_idx,
                format_func=fmt_month,
                key=f"su_act_{chosen_filing}",
                help="The month R&D activities took place.",
            )

        entity_changed = chosen_entity != draft.get("entity")
        filing_changed = chosen_filing != draft.get("reporting_month")
        act_changed    = chosen_act    != draft.get("activities_month")

        if entity_changed or filing_changed or act_changed:
            # Changing entity or filing month is NAVIGATION, not an edit: load the
            # report that actually lives at the new (entity, period) address.
            # Previously only the filing month triggered a reload, so switching
            # entity relabelled the current report's initiatives instead —
            # silently retagging Entity 107 work as Entity 108.
            if entity_changed or filing_changed:
                existing = load_submission(user, chosen_entity, chosen_filing)
                if existing:
                    draft = existing
                else:
                    draft = empty_draft()
                    draft["activities_month"] = prev_month_of(chosen_filing)
                draft["entity"]          = chosen_entity
                draft["reporting_month"] = chosen_filing
                st.session_state.draft = draft
                st.session_state.show_entry_picker = False
                st.rerun()

            # Reporting-period change is a real edit, so refuse it on a locked
            # report rather than mutating a submitted or archived record.
            elif act_changed:
                if is_locked(draft):
                    st.warning(
                        f"This report is {STATUS_LABELS.get(computed_status(draft), '')} "
                        "and can't be changed. Ask the Oversight Lead to reopen it."
                    )
                else:
                    draft["activities_month"] = chosen_act
                    save_draft(user, draft)
                    st.session_state.draft = draft
                    st.rerun()

        if setup_complete:
            fname = export_filename(draft["entity"], draft["reporting_month"])
            st.success(
                f"✓ **Filing Month:** {fmt_month(draft['reporting_month'])} "
                f"(Entity {draft['entity']}, filename: **{fname}**)  |  "
                f"**Reporting Period:** {fmt_month(draft.get('activities_month',''))}"
            )

    if not setup_complete:
        st.info("Complete the Report Setup above before adding initiatives.")
        return

    # ── Ask about carried-forward work before anything else ──────────────────
    # Each initiative rolled into this period needs one of the three pathways
    # before the report means anything. This used to sit behind ✏ Edit inside a
    # collapsed expander, so in practice it was never found and carried-forward
    # rows were resubmitted with last month's activities still in them.
    #
    # Deferring is allowed — being unable to reach the rest of the dashboard
    # would be worse than an unanswered prompt — but the deferral lasts only
    # for this session, and a banner keeps the outstanding count visible.
    if not locked:
        nxt = next_pathway_prompt(draft)
        if nxt is not None:
            st.session_state.wiz_init = dict(nxt)
            st.session_state.wiz_step = 0
            st.session_state.screen   = "pathway_select"
            st.rerun()

        deferred = pathways_pending(draft)
        if deferred:
            st.warning(
                f"⏳ {len(deferred)} carried-forward initiative"
                f"{'s' if len(deferred) != 1 else ''} still need"
                f"{'' if len(deferred) != 1 else 's'} a monthly update: "
                + ", ".join(f"**{i.get('initiative_name','Unnamed')}**" for i in deferred)
                + ". Open one below and choose what happened with it."
            )

    st.divider()

    # ── Carry-over from any past period ─────────────────────────────────────
    # Searches ALL past months for this user (not just previous month)
    # so nothing falls through the cracks.
    past_inits = get_past_ongoing_initiatives(user, draft["reporting_month"])

    # Filter out initiatives already in the current draft (by name)
    # Compare on series, not name — the same initiative may already be present as
    # a carried-forward row under a slightly different label.
    present_series = {series_id_of(i) for i in draft["initiatives"]}
    past_inits     = [i for i in past_inits if series_id_of(i) not in present_series]

    if not submitted and past_inits:
        st.markdown(f"""
        <div class="carryover-banner">
            <strong>📋 Ongoing initiatives from past periods</strong><br>
            {len(past_inits)} active initiative{"s" if len(past_inits)!=1 else ""} found across your history.
            Carry any forward — key details pre-fill; you only update activities and notes.
        </div>
        """, unsafe_allow_html=True)
        cols = st.columns(min(len(past_inits), 3))
        for idx, pi in enumerate(past_inits):
            with cols[idx % 3]:
                if st.button(f"↩ {pi.get('initiative_name','Unnamed')}", key=f"co_{pi['id']}"):
                    st.session_state.wiz_init = carryover_initiative(pi)
                    st.session_state.wiz_step = 0
                    st.session_state.wiz_mode = "carryover"
                    st.session_state.screen   = "wizard"
                    st.rerun()

    st.divider()

    # ── Status strip ─────────────────────────────────────────────────────────
    c1, c2, c3 = st.columns([4, 1, 1])
    with c1:
        st.markdown(f"**Report Status:** {badge_html(status)}", unsafe_allow_html=True)
        if draft.get("submitted_at"):
            st.caption(f"Submitted {ts_to_et(draft['submitted_at'])}")

        if status == "rejected":
            flagged = outstanding_returns(draft)
            total   = len(draft.get("initiatives") or [])
            st.warning(
                f"🔴 {len(flagged)} of {total} initiative"
                f"{'s' if total != 1 else ''} need"
                f"{'' if len(flagged) != 1 else 's'} revision. "
                "Everything else has been accepted — update the flagged "
                f"initiative{'s' if len(flagged) != 1 else ''} below and resubmit."
            )
            for note in outstanding_review_notes(draft):
                st.markdown(f"&nbsp;&nbsp;• {note}", unsafe_allow_html=True)
        elif status == "archived":
            arc_str = f" on {ts_to_et(draft['archived_at'], '%b %d, %Y')}" if draft.get("archived_at") else ""
            st.info(
                f"📦 Archived{arc_str}. This report is read-only. "
                "Ask the Oversight Lead to reopen it if it needs changing."
            )
        elif status == "submitted":
            if review_started(draft):
                st.info(
                    "🔵 The Oversight Lead has started reviewing this report, so it "
                    "can't be withdrawn. Anything they send back will show up here."
                )
            else:
                st.info("🔵 With the Oversight Lead for review — not looked at yet.")
                if st.button("✏ Make changes", key="unsubmit_btn",
                             help="Withdraws the submission so you can keep editing"):
                    unsubmit(draft)
                    save_draft(user, draft)
                    st.session_state.draft = draft
                    st.success("Withdrawn. Edit what you need, then submit again.")
                    st.rerun()
    with c2:
        st.metric("Initiatives", len(current_rows(draft)))
    with c3:
        if st.button("📦 Archive →", key="nav_archive",
                     help="Archived reports for the whole team, read-only"):
            st.session_state.screen = "archive"
            st.rerun()
        if st.session_state.is_admin and st.button("Admin View →"):
            st.session_state.screen = "admin"
            st.rerun()

    # ── Initiatives list ──────────────────────────────────────────────────────
    c1, c2 = st.columns([5, 1])
    with c1:
        act_m = draft.get("activities_month") or draft.get("reporting_month","")
        st.subheader(f"Initiatives — {fmt_month(act_m)}")
        st.caption(
            f"R&D activities that took place in **{fmt_month(act_m)}**. "
            "This is the period that will appear in the export."
        )
    with c2:
        if not submitted:
            existing_inits = draft.get("initiatives", [])
            n_existing = len(existing_inits)

            # Show context if there are already initiatives in this report
            if n_existing > 0:
                n_approved = sum(1 for i in existing_inits if i.get("initiative_status") == "approved")
                n_returned = sum(1 for i in existing_inits if i.get("initiative_status") == "returned")
                parts = [f"{n_existing} initiative{'s' if n_existing!=1 else ''} already in this report"]
                if n_approved:
                    parts.append(f"{n_approved} accepted by admin")
                if n_returned:
                    parts.append(f"⚠ {n_returned} need{'s' if n_returned==1 else ''} revision")
                st.caption(" · ".join(parts) + " — adding more will include them in the same submission.")

            if st.button("＋ Add Initiative", type="primary", width='stretch'):
                st.session_state.screen = "entry_picker"
                st.rerun()

    if not draft["initiatives"]:
        st.info("No initiatives added yet. Use **＋ Guided Entry** to add one at a time, or **📊 Bulk Entry** to fill a spreadsheet grid for multiple initiatives at once.")
    else:
        # ── Quick-scan summary table ────────────────────────────────────────
        # Lets you see everything at a glance before opening any single card —
        # this is what matters once there are more than a couple initiatives.
        status_icon_map = {"approved": "✅", "returned": "🔴", "active": "🔵"}
        # Only this period's rows are actionable. Rows carried in from earlier
        # months are a fixed record of what was reported then, and appear in the
        # read-only history block below.
        editable_inits = current_rows(draft)
        past_rows      = historical_rows(draft)

        table_rows = []
        for init in editable_inits:
            istatus = init.get("initiative_status", "active")
            table_rows.append({
                " ":                  status_icon_map.get(istatus, "🔵"),
                GROUP_LABEL:          init.get("employee_group", ""),
                "Initiative":         init.get("initiative_name", "Unnamed"),
                "Business Component": init.get("business_component", ""),
                "Team":               ", ".join(all_people_on(init)),
                "Start":              init.get("start_date", "—"),
                "End":                init.get("expected_end_date", "—"),
            })
        st.dataframe(
            table_rows,
            width='stretch',
            hide_index=True,
            column_config={" ": st.column_config.TextColumn(width="small")},
        )
        st.caption("✅ Accepted   ·   🔴 Rejected — needs revision   ·   🔵 Pending review")

        if past_rows:
            months = sorted({i.get("month_yr", "") for i in past_rows if i.get("month_yr")}, reverse=True)
            with st.expander(
                f"📖 Earlier months in this report ({len(past_rows)} row"
                f"{'s' if len(past_rows) != 1 else ''}) — read-only"
            ):
                st.caption(
                    "Carried forward when this period was rolled over. These are the "
                    "records of what was reported in "
                    + ", ".join(fmt_month(m) for m in months)
                    + ". They stay in the report and in the export, and can't be edited."
                )
                st.dataframe(
                    [{
                        "Month": fmt_month(i.get("month_yr", "")),
                        "Initiative": i.get("initiative_name", "Unnamed"),
                        "Activities": (i.get("activities", "") or "")[:120],
                        "Team": ", ".join(i.get("team_members") or []),
                    } for i in sorted(past_rows, key=lambda x: x.get("month_yr", ""), reverse=True)],
                    width='stretch',
                    hide_index=True,
                )

        st.write("")

        for init in editable_inits:
            iid = init["id"]
            istatus = init.get("initiative_status", "active")
            status_icon = status_icon_map.get(istatus, "🔵")
            with st.expander(
                f"{status_icon} {'↩ ' if init.get('carry_over') else ''}"
                f"{init.get('initiative_name','Unnamed')} — {init.get('business_component','')}",
                expanded=False,
            ):
                st.markdown(f"**{init.get('initiative_description','—')}**")
                st.caption(
                    f"📅 {init.get('start_date','—')} → {init.get('expected_end_date','—')}  "
                    f"  👥 {', '.join(all_people_on(init) or ['—'])}"
                )
                if init.get("tech_uncertainty"):
                    st.markdown("**Technical Uncertainty:**")
                    st.markdown(
                        f'<div style="background:#f8fafc;padding:10px 14px;border-radius:6px;'
                        f'font-size:13px;color:#334155;">{init["tech_uncertainty"]}</div>',
                        unsafe_allow_html=True,
                    )
                if init.get("activities"):
                    st.markdown("**Activities to eliminate uncertainty:**")
                    st.caption(init["activities"])
                if init.get("notes"):
                    st.markdown(f"*Notes: {init['notes']}*")

                # Reviewer feedback on this specific initiative
                if istatus == "returned":
                    ret_str = f" on {ts_to_et(init['returned_at'], '%b %d %I:%M %p')}" if init.get("returned_at") else ""
                    rc = (init.get("review_comment") or "").strip()
                    st.warning(
                        f"🔴 Revision requested{ret_str}."
                        + (f'  Reviewer note: "{rc}"' if rc else "")
                        + "  Edit this initiative, then resubmit the report."
                    )
                elif istatus == "approved":
                    appr_str = f" on {ts_to_et(init['approved_at'], '%b %d %I:%M %p')}" if init.get("approved_at") else ""
                    st.success(f"✓ Accepted by the Oversight Lead{appr_str}.")

                # Earlier rounds of feedback, kept for context once addressed.
                if init.get("review_history"):
                    with st.expander(f"Previous feedback ({len(init['review_history'])})"):
                        for h in init["review_history"]:
                            when = ts_to_et(h.get("at"), "%b %d, %Y") if h.get("at") else ""
                            st.caption(f"{when} — {h.get('comment','')}")

                st.write("")
                # Editing is blocked whenever the report is locked, not only when
                # archived. Deleting from a submitted or approved report used to
                # silently revert it to In Progress, un-approving work the
                # Oversight Lead had already signed off without telling anyone.
                lock_help = (
                    f"This report is {STATUS_LABELS.get(status, status)} and can't be edited."
                    if locked else None
                )
                c1, c2, c3 = st.columns([4, 0.8, 0.8])
                with c2:
                    if st.button("✏ Edit", key=f"edit_{iid}", disabled=locked, help=lock_help):
                        st.session_state.wiz_init = dict(init)
                        st.session_state.wiz_step = 0
                        if init.get("carry_over"):
                            st.session_state.screen = "pathway_select"
                        else:
                            st.session_state.wiz_mode = "edit"
                            st.session_state.screen   = "wizard"
                        st.rerun()
                with c3:
                    if st.button("🗑 Delete", key=f"del_{iid}", disabled=locked, help=lock_help):
                        st.session_state.confirm_del = iid
                        st.rerun()

                if st.session_state.get("confirm_del") == iid:
                    st.markdown('<div class="delete-confirm">', unsafe_allow_html=True)
                    st.warning(
                        f"Delete **{init.get('initiative_name','this initiative')}**? "
                        "This cannot be undone."
                    )
                    ca, cb = st.columns(2)
                    with ca:
                        if st.button("Yes, delete it", key=f"conf_yes_{iid}", type="primary"):
                            draft["initiatives"] = [i for i in draft["initiatives"] if i["id"] != iid]
                            if not current_rows(draft):
                                # An empty report was never really submitted.
                                draft.pop("submitted_at", None)
                            save_draft(user, draft)
                            st.session_state.draft       = draft
                            st.session_state.confirm_del = None
                            st.success("Initiative deleted.")
                            st.rerun()
                    with cb:
                        if st.button("Cancel", key=f"conf_no_{iid}"):
                            st.session_state.confirm_del = None
                            st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

    # ── Submit + Export ───────────────────────────────────────────────────────
    st.divider()
    c1, c2 = st.columns([3, 1])
    with c1:
        if not locked and draft["initiatives"]:
            revising = status == "rejected"
            n = len(draft["initiatives"])
            st.markdown("#### Ready to resubmit?" if revising else "#### Ready to submit?")
            st.caption(
                "Sends your revised report back to the Oversight Lead for review."
                if revising else
                f"Sends {n} initiative{'s' if n != 1 else ''} to the Oversight Lead for review."
            )
            if st.button(
                "Resubmit for review" if revising else "Submit for review",
                type="primary",
            ):
                mark_submitted(draft)
                save_draft(user, draft)
                st.session_state.draft = draft
                st.success("Resubmitted for review." if revising else "Submitted for review.")
                st.rerun()
    with c2:
        if draft.get("initiatives"):
            xlsx = build_excel_individual(user, draft)
            fname = export_filename(draft.get("entity",""), draft.get("reporting_month",""))
            st.download_button(
                "↓ Download This Filing",
                data=xlsx,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Downloads this period's report only. Use Export My Reports below to download across multiple periods.",
            )

    # ── Export ────────────────────────────────────────────────────────────────
    st.write("")
    render_user_export_section(user)

    # ── History ──────────────────────────────────────────────────────────────
    st.write("")
    render_history_section(
        user, draft.get("reporting_month", ""), draft.get("entity", "")
    )


# ── User Export ───────────────────────────────────────────────────────────────

def render_user_export_section(user: str):
    """Filter and download this user's own reports as Excel."""
    reports = user_reports(user)
    if not reports:
        return

    with st.expander("⬇ Export my reports", expanded=False):
        st.caption("Filter your reports and download them as one Excel file.")
        st.write("")

        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            entities = sorted({r.get("entity", "") for r in reports if r.get("entity")})
            f_entity = st.selectbox("Entity", ["All"] + entities, key="ue_entity")
        with fc2:
            filings = sorted({r.get("reporting_month", "") for r in reports}, reverse=True)
            f_filing = st.selectbox(
                "Filing month", ["All"] + filings, format_func=lambda m: m if m == "All" else fmt_month(m),
                key="ue_filing",
            )
        with fc3:
            periods = sorted(
                {r.get("activities_month", "") for r in reports if r.get("activities_month")},
                reverse=True,
            )
            f_period = st.selectbox(
                "Reporting period", ["All"] + periods,
                format_func=lambda m: m if m == "All" else fmt_month(m),
                key="ue_period",
            )

        st.write("")
        st.caption("**Status**")
        present = sorted({computed_status(r) for r in reports})
        chosen_statuses = []
        cols = st.columns(max(len(present), 1))
        for i, key in enumerate(present):
            with cols[i]:
                if st.checkbox(STATUS_LABELS.get(key, key), value=True, key=f"ue_status_{key}"):
                    chosen_statuses.append(key)

        filtered = [
            r for r in reports
            if (f_entity == "All" or r.get("entity") == f_entity)
            and (f_filing == "All" or r.get("reporting_month") == f_filing)
            and (f_period == "All" or r.get("activities_month") == f_period)
            and computed_status(r) in chosen_statuses
        ]

        st.write("")
        if not filtered:
            st.caption("No reports match these filters. Widen a filter to see results.")
            return

        n_inits = sum(len(r.get("initiatives") or []) for r in filtered)
        st.caption(
            f"{len(filtered)} report{'s' if len(filtered) != 1 else ''} · "
            f"{n_inits} initiative{'s' if n_inits != 1 else ''}"
        )

        if len(filtered) == 1:
            rep = filtered[0]
            st.caption("One period matches, so this downloads that single filing.")
            st.download_button(
                f"↓ Download — Entity {rep.get('entity','')} {fmt_month(rep.get('reporting_month',''))}",
                data=build_excel_individual(user, rep),
                file_name=export_filename(rep.get("entity", ""), rep.get("reporting_month", "")),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="ue_dl_single",
            )
        else:
            st.download_button(
                f"↓ Download {len(filtered)} reports",
                data=build_excel_consolidated(
                    filtered, group_by="entity", status_desc=status_desc_for(chosen_statuses)
                ),
                file_name=f"{_safe_name(user)}_RD_Report_{datetime.now().strftime('%m%d%y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="ue_dl_multi",
            )


# ── History (shown below submit on the dashboard) ───────────────────────────

def render_history_section(user: str, current_reporting_month: str, current_entity: str = ""):
    """This user's own past reports. The Archive tab shows everyone's."""
    history = [
        r for r in get_user_history(user)
        if not (
            r.get("reporting_month") == current_reporting_month
            and r.get("entity") == current_entity
        )
    ]
    if not history:
        return

    with st.expander(
        f"📁 My submission history ({len(history)} report{'s' if len(history) != 1 else ''})",
        expanded=False,
    ):
        for rep in history:
            entity = rep.get("entity", "—")
            month  = rep.get("reporting_month", "")
            inits  = rep.get("initiatives") or []
            st.markdown(
                f"**Entity {entity} — {fmt_month(month)}** &nbsp; "
                f"{badge_html(computed_status(rep))} &nbsp; "
                f"*{len(inits)} initiative{'s' if len(inits) != 1 else ''}*",
                unsafe_allow_html=True,
            )
            for i in inits:
                st.markdown(
                    f"&nbsp;&nbsp;&nbsp;• **{i.get('initiative_name','Unnamed')}** — "
                    f"{i.get('business_component','')}  "
                    f"📅 {i.get('start_date','—')} → {i.get('expected_end_date','—')}  "
                    f"👥 {', '.join(all_people_on(i) or ['—'])}",
                    unsafe_allow_html=True,
                )
            if inits:
                st.download_button(
                    f"↓ Entity {entity} — {fmt_month(month)}",
                    data=build_excel_individual(user, rep),
                    file_name=export_filename(entity, month),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"hist_dl_{entity}_{month}",
                )
            st.write("")


# ── Wizard ────────────────────────────────────────────────────────────────────

def screen_wizard():
    user  = st.session_state.user
    init  = st.session_state.wiz_init
    step  = st.session_state.wiz_step
    mode  = st.session_state.wiz_mode
    s     = WIZARD_STEPS[step]
    total = len(WIZARD_STEPS)
    pct   = int((step+1)/total*100)

    is_carryover = mode == "carryover"
    is_prefilled = is_carryover and s["field"] in CARRYOVER_FIELDS

    c1, c2 = st.columns([5, 1])
    with c1:
        title = {"new":"New Initiative","edit":"Edit Initiative","carryover":"Carry Over Initiative"}[mode]
        st.markdown(f"### {title}")
    with c2:
        if st.button("✕ Cancel"):
            st.session_state.screen = "dashboard"
            st.rerun()

    st.markdown(f"""
    <div class="progress-bar-outer">
        <div class="progress-bar-inner" style="width:{pct}%"></div>
    </div>
    <p style="font-size:12px;color:#64748b;margin-top:2px;">
        Step {step+1} of {total} &nbsp;·&nbsp; {pct}% complete
    </p>
    """, unsafe_allow_html=True)

    if is_carryover and not is_prefilled:
        st.markdown(f"""<div class="carryover-banner">
            <strong>Updating for {fmt_month(st.session_state.draft.get('reporting_month',''))}:</strong>
            {init.get("initiative_name","this initiative")}
        </div>""", unsafe_allow_html=True)
    elif is_prefilled:
        st.markdown("""<div class="prefilled-banner">
            Pre-filled from last month — confirm or update before continuing.
        </div>""", unsafe_allow_html=True)

    st.markdown(f'<p class="step-label">Question {step+1} of {total}</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="wizard-question">{s["question"]}</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="wizard-hint">{s["hint"]}</p>', unsafe_allow_html=True)

    field   = s["field"]
    val     = init.get(field)
    new_val = val

    if s["type"] == "text":
        new_val = st.text_input(s["label"], value=val or "", placeholder=s.get("placeholder",""))
    elif s["type"] == "textarea":
        new_val = st.text_area(s["label"], value=val or "", placeholder=s.get("placeholder",""), height=140)
    elif s["type"] == "date":
        parsed = None
        if val:
            try:
                parsed = datetime.strptime(str(val), "%Y-%m-%d").date()
            except Exception:
                pass
        picked  = st.date_input(s["label"], value=parsed)
        new_val = picked.strftime("%Y-%m-%d") if picked else None
    elif s["type"] == "group":
        new_val = group_select(f"wiz_grp_{init.get('id','new')}", val or "")
        # Team members belong to a group, so changing it invalidates them.
        # Say so and drop them rather than exporting a team that contradicts
        # the group on the same row.
        if new_val and val and new_val != val:
            kept = _prune_to_group(init.get("team_members"), new_val)
            dropped = [m for m in (init.get("team_members") or []) if m not in kept]
            if dropped:
                st.warning(
                    f"{len(dropped)} team member"
                    f"{'s are' if len(dropped) != 1 else ' is'} not in "
                    f"{new_val} and will be removed: {', '.join(dropped)}."
                )
                init["team_members"] = kept
                clear_team_picker_state(f"wiz_team_{init.get('id','new')}")
    elif s["type"] == "multiselect":
        # Scoped to the group chosen in step 1. Keyed on the initiative so each
        # keeps its own selection across reruns, and seeded from the saved value
        # so edits pre-fill.
        split_legacy_members(init)
        new_val, cons = render_team_picker(
            f"wiz_team_{init.get('id','new')}",
            init.get("employee_group", ""),
            val or [],
            label=s["label"],
            current_contractors=init.get("contractors") or [],
        )
        init["contractors"] = cons

    init[field] = new_val

    # A closed business component must not be reused — that is the whole point
    # of closing it. Checked here rather than only on save so the preparer finds
    # out at the field, not after answering nine more questions.
    bc_block = ""
    if field == "business_component" and str(new_val or "").strip():
        rec = bc_is_closed(st.session_state.draft.get("entity", ""), str(new_val))
        if rec:
            when = ts_to_et(rec.get("closed_at"), "%b %d, %Y") if rec.get("closed_at") else ""
            bc_block = (
                f"**{rec.get('label', new_val)}** was closed"
                + (f" by {rec.get('closed_by')}" if rec.get("closed_by") else "")
                + (f" on {when}" if when else "")
                + ". Use a different business component, or ask the Oversight "
                  "Lead to reopen it in Admin → Settings."
            )
            st.error(bc_block)

    if s["required"]:
        if s["type"] == "multiselect":
            # A one-person contractor team is a real answer, so don't insist on
            # a directory employee being named.
            is_valid = bool(new_val) or bool(init.get("contractors"))
        elif s["type"] == "group":
            # Without a directory there is nothing to choose from, so don't
            # block the preparer on a field the admin hasn't set up yet.
            is_valid = bool(str(new_val or "").strip()) or not directory.is_loaded()
        else:
            is_valid = bool(str(new_val or "").strip())
    else:
        is_valid = True

    if bc_block:
        is_valid = False

    c1, c2, c3 = st.columns([1, 4, 1])
    with c1:
        if st.button("← Back", disabled=(step == 0)):
            st.session_state.wiz_step -= 1
            st.rerun()
    with c3:
        is_last = step == total - 1
        if st.button("Save Initiative ✓" if is_last else "Next →", disabled=not is_valid, type="primary"):
            if is_last:
                draft = st.session_state.draft
                # Stamp month_yr on new/carryover initiatives so the export
                # always shows the correct period regardless of later rollovers.
                # Edit mode preserves whatever month_yr was already on the initiative.
                if mode in ("new", "carryover") and not init.get("month_yr"):
                    # Use activities_month (the R&D period) not filing month
                    init["month_yr"] = draft.get("activities_month") or draft.get("reporting_month", "")
                init.setdefault("historical", False)
                init["series_id"] = series_id_of(init)
                if mode == "edit":
                    clear_return_flag(init)
                    draft["initiatives"] = [i if i["id"] != init["id"] else init for i in draft["initiatives"]]
                else:
                    draft["initiatives"].append(init)
                save_draft(user, draft)
                st.session_state.draft  = draft
                st.session_state.screen = "dashboard"
            else:
                st.session_state.wiz_init = init
                st.session_state.wiz_step += 1
            st.rerun()

    dots = "".join(
        f'<span style="display:inline-block;width:{"24px" if i==step else "8px"};height:8px;'
        f'border-radius:4px;margin:0 3px;'
        f'background:{"#c86a2a" if i<step else "#1a3c5e" if i==step else "#cbd5e1"}"></span>'
        for i in range(total)
    )
    st.markdown(f'<div style="text-align:center;margin-top:20px;">{dots}</div>', unsafe_allow_html=True)


# ── Admin ─────────────────────────────────────────────────────────────────────

def render_archive_browser(key_prefix: str = "arc"):
    """Read-only view of every archived report, for everyone.

    Deliberately not filtered by the entity permissions in Settings — those
    govern which entities a person may FILE for. The archive is a shared record
    of closed periods that the whole team can read and export.

    No edit controls are rendered here at all, rather than rendered-and-disabled,
    so there is no path from this screen to a write.
    """
    archived = all_reports(statuses=["archived"])

    if not archived:
        st.info(
            "Nothing archived yet. Reports appear here once the Oversight Lead "
            "archives them from the Submissions tab."
        )
        return

    st.caption(
        f"{len(archived)} archived report"
        f"{'s' if len(archived) != 1 else ''} across the whole team. "
        "Read-only — archived reports can't be edited by anyone until the "
        "Oversight Lead reopens them."
    )
    st.write("")

    f1, f2, f3 = st.columns(3)
    with f1:
        entities = sorted({r.get("entity", "") for r in archived if r.get("entity")})
        f_entity = st.selectbox("Entity", ["All"] + entities, key=f"{key_prefix}_entity")
    with f2:
        periods = sorted({r.get("reporting_month", "") for r in archived}, reverse=True)
        f_period = st.selectbox(
            "Filing month", ["All"] + periods,
            format_func=lambda m: m if m == "All" else fmt_month(m),
            key=f"{key_prefix}_period",
        )
    with f3:
        people = sorted({r.get("user", "") for r in archived if r.get("user")})
        f_user = st.selectbox("Team member", ["All"] + people, key=f"{key_prefix}_user")

    shown = [
        r for r in archived
        if (f_entity == "All" or r.get("entity") == f_entity)
        and (f_period == "All" or r.get("reporting_month") == f_period)
        and (f_user   == "All" or r.get("user") == f_user)
    ]

    st.write("")
    if not shown:
        st.caption("No archived reports match these filters.")
        return

    unique_rows, dropped = dedupe_stats(shown)
    m1, m2, m3 = st.columns(3)
    m1.metric("Reports", len(shown))
    m2.metric("Rows", unique_rows)
    m3.metric("Team members", len({r.get("user") for r in shown}))
    if dropped:
        st.caption(
            f"ℹ {dropped} duplicate row{'s' if dropped != 1 else ''} collapsed in the "
            "download — later periods carry copies of earlier months."
        )

    st.download_button(
        f"↓ Download {len(shown)} archived report{'s' if len(shown) != 1 else ''}",
        data=build_excel_consolidated(shown, group_by="entity", status_desc="Archived"),
        file_name=f"RD_Archive_{datetime.now().strftime('%m%d%y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"{key_prefix}_dl_all",
    )

    st.write("")
    st.divider()

    for rep in shown:
        entity = rep.get("entity", "—")
        month  = rep.get("reporting_month", "")
        who    = rep.get("user", "Unknown")
        inits  = rep.get("initiatives") or []
        arc    = f" · archived {ts_to_et(rep['archived_at'], '%b %d, %Y')}" if rep.get("archived_at") else ""

        with st.expander(
            f"📦 {who} — Entity {entity} — {fmt_month(month)} "
            f"({len(inits)} initiative{'s' if len(inits) != 1 else ''})"
        ):
            st.caption(
                f"Reporting period: {fmt_month(rep.get('activities_month','')) or '—'}{arc}"
            )
            for i in inits:
                st.markdown(f"**{i.get('initiative_name','Unnamed')}** — {i.get('business_component','')}")
                st.caption(
                    f"📅 {i.get('start_date','—')} → {i.get('expected_end_date','—')}  ·  "
                    f"👥 {', '.join(all_people_on(i) or ['—'])}"
                )
                if i.get("initiative_description"):
                    st.markdown(i["initiative_description"])
                if i.get("tech_uncertainty"):
                    st.markdown("*Technical uncertainty:*")
                    st.caption(i["tech_uncertainty"])
                if i.get("activities"):
                    st.markdown("*Activities:*")
                    st.caption(i["activities"])
                if i.get("notes"):
                    st.caption(f"Notes: {i['notes']}")
                st.write("")

            st.download_button(
                f"↓ {who} — Entity {entity} {fmt_month(month)}",
                data=build_excel_individual(who, rep),
                file_name=export_filename(entity, month).replace(
                    ".xlsx", f"_{_safe_name(who)}.xlsx"
                ),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{key_prefix}_dl_{_safe_name(who)}_{entity}_{month}",
            )


def screen_archive():
    """Standalone archive screen for preparers."""
    c1, c2 = st.columns([5, 1])
    with c1:
        st.markdown("## 📦 Archive")
        st.caption("Every archived report across the team. Read-only.")
    with c2:
        st.write("")
        if st.button("← Back"):
            st.session_state.screen = "dashboard"
            st.rerun()
    st.divider()
    render_archive_browser("arcuser")


# ── Admin ─────────────────────────────────────────────────────────────────────

def _select_all_controls(scope_key: str, keys: list[str], label: str):
    """Select-all / clear-all as BUTTONS, not a parent checkbox.

    A parent checkbox forced its children true on every rerun, which meant an
    individual month could never be unticked while the parent was on. Buttons set
    state once and then leave the child checkboxes alone.
    """
    b1, b2 = st.columns(2)
    with b1:
        if st.button(f"Select all {label}", key=f"{scope_key}_all"):
            for k in keys:
                st.session_state[k] = True
            st.rerun()
    with b2:
        if st.button(f"Clear all {label}", key=f"{scope_key}_none"):
            for k in keys:
                st.session_state[k] = False
            st.rerun()


def screen_admin():
    st.markdown("## ⚙ Admin Dashboard")

    c1, c2 = st.columns([5, 1])
    with c2:
        if st.button("← Back"):
            st.session_state.screen = "login" if st.session_state.user == "Admin" else "dashboard"
            st.rerun()

    reports = all_reports()
    combos  = get_combos()

    render_admin_reminder()

    # ── Summary ───────────────────────────────────────────────────────────────
    by_status = {}
    for r in reports:
        by_status.setdefault(computed_status(r), []).append(r)

    n_inits = sum(len(r.get("initiatives") or []) for r in reports)
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Team members",   len({r.get("user") for r in reports}))
    c2.metric("Reports",        len(reports))
    c3.metric("Awaiting review", len(by_status.get("submitted", [])))
    c4.metric("Approved",       len(by_status.get("approved", [])))
    c5.metric("Archived",       len(by_status.get("archived", [])))
    c6.metric("Initiatives",    n_inits)

    st.write("")

    ADMIN_TABS = [
        "📁 Submissions", "🔄 Rollover", "📦 Archive",
        "⬇ Export", "🗄 Backup & Data", "⚙ Settings",
    ]
    if "admin_tab" not in st.session_state or st.session_state.admin_tab not in ADMIN_TABS:
        st.session_state.admin_tab = ADMIN_TABS[0]
    active_tab = st.radio(
        "##nav", ADMIN_TABS,
        index=ADMIN_TABS.index(st.session_state.admin_tab),
        horizontal=True, label_visibility="collapsed", key="admin_tab_radio",
    )
    st.session_state.admin_tab = active_tab
    st.write("")

    # ══ Submissions ═══════════════════════════════════════════════════════════
    if active_tab == "📁 Submissions":
        if not reports:
            st.info("No reports yet. They appear here once a team member submits one.")
        else:
            st.caption(
                "**Filing month** is when the report was submitted — it sets the Group "
                "number in the filename. **Reporting period** is when the R&D work "
                "happened, shown inside each report."
            )
            fc1, fc2, fc3 = st.columns([1.2, 1.5, 1.5])
            with fc1:
                entities = sorted({e for e, _ in combos})
                f_entity = st.selectbox("Entity", ["All"] + entities, key="f_entity")
            with fc2:
                periods = sorted({rm for _, rm in combos}, reverse=True)
                f_period = st.selectbox(
                    "Filing month", ["All"] + periods,
                    format_func=lambda m: m if m == "All" else fmt_month(m),
                    key="f_period",
                )
            with fc3:
                status_opts = ["All"] + [STATUS_LABELS[k] for k in
                    ("submitted", "rejected", "approved", "archived", "in-progress")]
                f_status_label = st.selectbox("Status", status_opts, key="f_status")

            label_to_key = {v: k for k, v in STATUS_LABELS.items()}
            f_status = label_to_key.get(f_status_label)

            shown = [
                r for r in reports
                if (f_entity == "All" or r.get("entity") == f_entity)
                and (f_period == "All" or r.get("reporting_month") == f_period)
                and (f_status is None or computed_status(r) == f_status)
            ]

            if not shown:
                st.caption("No reports match these filters.")

            # Group by (entity, period) for readability
            groups: dict[tuple[str, str], list[dict]] = {}
            for r in shown:
                groups.setdefault((r.get("entity", ""), r.get("reporting_month", "")), []).append(r)

            for (entity, rm) in sorted(groups.keys(), key=lambda c: (c[1], c[0]), reverse=True):
                group = sorted(groups[(entity, rm)], key=lambda r: r.get("user", ""))
                sample_am = next((r.get("activities_month") for r in group if r.get("activities_month")), None)
                am_suffix = f"  ·  Reporting period: **{fmt_month(sample_am)}**" if sample_am else ""
                st.markdown(f"#### Entity {entity} — Filing month: **{fmt_month(rm)}**{am_suffix}")

                for rep in group:
                    username = rep.get("user", "Unknown")
                    rstatus  = computed_status(rep)
                    inits    = rep.get("initiatives") or []
                    n_cur    = len(current_rows(rep))
                    n_past   = len(historical_rows(rep))
                    icon     = {"approved": "✅", "submitted": "🔵", "in-progress": "🟡",
                                "rejected": "🔴", "archived": "📦"}.get(rstatus, "⚪")
                    rkey     = f"{entity}_{rm}_{_safe_name(username)}"

                    with st.expander(
                        f"{icon}  {username}   —  {STATUS_LABELS.get(rstatus, '—')}  "
                        f"({n_cur} this period"
                        + (f" · {n_past} carried" if n_past else "") + ")",
                        expanded=(rstatus == "submitted"),
                    ):
                        h1, h2 = st.columns([4, 1])
                        with h1:
                            st.markdown(
                                f"**Filing month:** {fmt_month(rm)} &nbsp;|&nbsp; "
                                f"**Reporting period:** {fmt_month(rep.get('activities_month','')) or '—'}",
                                unsafe_allow_html=True,
                            )
                            if rep.get("submitted_at"):
                                st.caption(f"Submitted: {ts_to_et(rep['submitted_at'], '%b %d %I:%M %p')}")
                            if rep.get("approved_at"):
                                st.caption(f"Approved: {ts_to_et(rep['approved_at'], '%b %d %I:%M %p')}")
                        with h2:
                            if inits:
                                st.download_button(
                                    "↓ Export",
                                    data=build_excel_individual(username, rep),
                                    file_name=export_filename(entity, rm).replace(
                                        ".xlsx", f"_{_safe_name(username)}.xlsx"
                                    ),
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"dl_{rkey}",
                                )

                        # ── Archived reports are read-only, full stop ─────────
                        if rstatus == "archived":
                            arc = ts_to_et(rep["archived_at"], "%b %d, %Y %I:%M %p") if rep.get("archived_at") else ""
                            st.info(
                                f"📦 Archived{f' on {arc}' if arc else ''}. "
                                "Read-only, and still included in every export."
                            )
                            if st.button("↩ Reopen for editing", key=f"reopen_{rkey}"):
                                # Reopening returns the report to whatever its
                                # initiatives say it is. There was previously no
                                # way back out of archived at all, despite the
                                # user-facing banner telling people to ask.
                                rep["status"] = "approved"
                                rep.pop("archived_at", None)
                                save_draft(username, rep)
                                st.success(f"Reopened {username}'s {fmt_month(rm)} report.")
                                st.rerun()

                            st.divider()
                            for i in sorted(inits, key=lambda x: x.get("month_yr", "")):
                                st.markdown(
                                    f"**{fmt_month(i.get('month_yr','')) or '—'}** · "
                                    f"**{i.get('initiative_name','Unnamed')}** — "
                                    f"{i.get('business_component','')}  \n"
                                    f"<small>👥 {', '.join(all_people_on(i) or ['—'])} "
                                    f"&nbsp;|&nbsp; 📅 {i.get('start_date','—')} → "
                                    f"{i.get('expected_end_date','—')}</small>",
                                    unsafe_allow_html=True,
                                )
                            continue

                        # ── Report-level actions ──────────────────────────────
                        if rstatus == "submitted":
                            ra1, ra2, ra3 = st.columns([3, 1, 1])
                            with ra2:
                                if st.button("✓ Accept all", key=f"appr_rpt_{rkey}", type="primary"):
                                    # Accepting archives this period and creates the
                                    # next one in a single step, the way Flow 4 was
                                    # triggered by the status change to Complete.
                                    did, tgt, why = accept_report(rep)
                                    if did:
                                        st.success(
                                            f"Accepted {username}'s report. "
                                            f"{fmt_month(rm)} is archived and "
                                            f"{fmt_month(tgt)} is waiting for them."
                                        )
                                    else:
                                        st.success(
                                            f"Accepted and archived {username}'s report. "
                                            f"Not rolled forward — {why}."
                                        )
                                    st.rerun()
                            with ra3:
                                if st.button("✕ Return all", key=f"rej_rpt_{rkey}"):
                                    st.session_state[f"return_report_{rkey}"] = True
                                    st.rerun()

                        if st.session_state.get(f"return_report_{rkey}"):
                            note = st.text_area(
                                "What needs to change? This note is shown to the preparer.",
                                key=f"rc_rpt_{rkey}", height=80,
                                placeholder="Describe what needs updating...",
                            )
                            rc1, rc2 = st.columns(2)
                            with rc1:
                                if st.button("Return for revision", key=f"rc_confirm_{rkey}", type="primary"):
                                    if not note.strip():
                                        st.error("Add a note so the preparer knows what to change.")
                                    else:
                                        now = int(time.time() * 1000)
                                        for i in current_rows(rep):
                                            i["initiative_status"] = "returned"
                                            i["returned_at"]       = now
                                            i["review_comment"]    = note.strip()
                                            i.pop("approved_at", None)
                                        st.session_state.pop(f"return_report_{rkey}", None)
                                        save_draft(username, rep)
                                        st.rerun()
                            with rc2:
                                if st.button("Cancel", key=f"rc_cancel_{rkey}"):
                                    st.session_state.pop(f"return_report_{rkey}", None)
                                    st.rerun()

                        if rstatus == "approved":
                            arc1, arc2 = st.columns([4, 1])
                            with arc1:
                                st.caption(
                                    "Accepted but not archived — usually because it was "
                                    "reopened. Archive it to close the period out."
                                )
                            with arc2:
                                if st.button("📦 Archive", key=f"archive_{rkey}"):
                                    rep["status"]      = "archived"
                                    rep["archived_at"] = int(time.time() * 1000)
                                    save_draft(username, rep)
                                    st.success("Archived.")
                                    st.rerun()

                        if rstatus == "rejected":
                            n = len(outstanding_returns(rep))
                            st.warning(
                                f"🔴 Waiting on the preparer — {n} initiative"
                                f"{'s' if n != 1 else ''} returned for revision."
                            )

                        st.divider()

                        if not inits:
                            st.caption("No initiatives.")
                        elif rstatus == "in-progress":
                            st.info(
                                "🟡 Not submitted yet. The preparer needs to review their "
                                "initiatives and submit for review before you can accept "
                                "or return them."
                            )
                            for i in current_rows(rep):
                                st.markdown(
                                    f"🔵 **{'↩ ' if i.get('carry_over') else ''}"
                                    f"{i.get('initiative_name','Unnamed')}** — "
                                    f"{i.get('business_component','')}  \n"
                                    f"<small>👥 {', '.join(all_people_on(i) or ['—'])} "
                                    f"&nbsp;|&nbsp; 📅 {i.get('start_date','—')} → "
                                    f"{i.get('expected_end_date','—')}</small>",
                                    unsafe_allow_html=True,
                                )
                        else:
                            past = historical_rows(rep)
                            if past:
                                months = sorted(
                                    {i.get("month_yr", "") for i in past if i.get("month_yr")},
                                    reverse=True,
                                )
                                with st.expander(
                                    f"📖 Earlier months carried into this report "
                                    f"({len(past)} row{'s' if len(past) != 1 else ''}) — read-only"
                                ):
                                    st.caption(
                                        "Already accepted in "
                                        + ", ".join(fmt_month(m) for m in months)
                                        + ". Kept for the record; nothing to review here."
                                    )
                                    st.dataframe(
                                        [{
                                            "Month": fmt_month(i.get("month_yr", "")),
                                            "Initiative": i.get("initiative_name", "Unnamed"),
                                            "Activities": (i.get("activities", "") or "")[:120],
                                        } for i in sorted(past, key=lambda x: x.get("month_yr", ""), reverse=True)],
                                        width='stretch', hide_index=True,
                                    )
                                st.write("")

                            for init in current_rows(rep):
                                iid     = init["id"]
                                istatus = init.get("initiative_status", "active")
                                iname   = init.get("initiative_name", "Unnamed")
                                ico     = {"approved": "✅", "returned": "🔴", "active": "🔵"}.get(istatus, "🔵")
                                ikey    = f"{rkey}_{iid}"

                                ic1, ic2, ic3 = st.columns([4, 0.9, 0.9])
                                with ic1:
                                    st.markdown(
                                        f"{ico} **{'↩ ' if init.get('carry_over') else ''}{iname}** — "
                                        f"{init.get('business_component','')}  \n"
                                        f"<small>👥 {', '.join(all_people_on(init) or ['—'])} "
                                        f"&nbsp;|&nbsp; 📅 {init.get('start_date','—')} → "
                                        f"{init.get('expected_end_date','—')}</small>",
                                        unsafe_allow_html=True,
                                    )
                                    if init.get("approved_at"):
                                        st.caption(f"   Accepted {ts_to_et(init['approved_at'], '%b %d %I:%M %p')}")
                                    if istatus == "returned" and init.get("returned_at"):
                                        rc = (init.get("review_comment") or "").strip()
                                        st.caption(
                                            f"   Returned {ts_to_et(init['returned_at'], '%b %d %I:%M %p')}"
                                            + (f' — "{rc}"' if rc else "")
                                        )
                                    if init.get("review_history"):
                                        st.caption(
                                            f"   {len(init['review_history'])} earlier "
                                            f"note{'s' if len(init['review_history']) != 1 else ''}, "
                                            "since addressed"
                                        )

                                with ic2:
                                    if istatus != "approved":
                                        if st.button("✓ Accept", key=f"appr_i_{ikey}", type="primary"):
                                            now = int(time.time() * 1000)
                                            clear_return_flag(init, reason="accepted")
                                            init["initiative_status"] = "approved"
                                            init["approved_at"]       = now
                                            save_draft(username, rep)
                                            # Accepting the LAST outstanding initiative
                                            # completes the report, so the same trigger
                                            # fires here as on Accept all — otherwise the
                                            # two review paths would behave differently.
                                            if computed_status(rep) == "approved":
                                                did, tgt, why = accept_report(rep)
                                                if did:
                                                    st.success(
                                                        f"Report complete. {fmt_month(rm)} archived, "
                                                        f"{fmt_month(tgt)} created for {username}."
                                                    )
                                                else:
                                                    st.success(
                                                        f"Report complete and archived. "
                                                        f"Not rolled forward — {why}."
                                                    )
                                            st.rerun()

                                with ic3:
                                    if istatus != "returned":
                                        if st.button("✕ Return", key=f"ret_i_{ikey}"):
                                            st.session_state[f"return_init_{ikey}"] = True
                                            st.rerun()

                                if st.session_state.get(f"return_init_{ikey}"):
                                    note_i = st.text_area(
                                        f'What needs to change on "{iname}"? '
                                        "This note is shown to the preparer.",
                                        key=f"rc_init_{ikey}", height=80,
                                        placeholder="Describe what needs updating...",
                                    )
                                    ri1, ri2 = st.columns(2)
                                    with ri1:
                                        if st.button("Return this initiative", key=f"rc_conf_i_{ikey}", type="primary"):
                                            if not note_i.strip():
                                                st.error("Add a note so the preparer knows what to change.")
                                            else:
                                                init["initiative_status"] = "returned"
                                                init["returned_at"]       = int(time.time() * 1000)
                                                init["review_comment"]    = note_i.strip()
                                                init.pop("approved_at", None)
                                                st.session_state.pop(f"return_init_{ikey}", None)
                                                # Returning ONE initiative no longer rejects the
                                                # whole report — status is derived, so everything
                                                # already accepted stays accepted.
                                                save_draft(username, rep)
                                                st.rerun()
                                    with ri2:
                                        if st.button("Cancel", key=f"rc_cancel_i_{ikey}"):
                                            st.session_state.pop(f"return_init_{ikey}", None)
                                            st.rerun()

                                st.write("")

                st.write("")

    # ══ Rollover ══════════════════════════════════════════════════════════════
    if active_tab == "🔄 Rollover":
        st.caption(
            "Accepting a report already archives it and creates the next period, so "
            "you don't normally need this tab. Use it to catch up after a gap, or to "
            "send a period to a month other than the next one. Anyone who already has "
            "data in the target month is skipped."
        )

        eligible = sorted(
            {
                (r.get("entity", ""), r.get("reporting_month", ""))
                for r in reports
                if computed_status(r) in ("approved", "archived")
            },
            reverse=True,
        )

        if not eligible:
            st.info(
                "Nothing to roll forward yet. Accept a report in the Submissions "
                "tab first."
            )
        else:
            chk_keys = [f"ro_chk_{e}_{m}" for e, m in eligible]
            _select_all_controls("ro", chk_keys, "periods")
            st.write("")

            selections = []
            for entity in sorted({e for e, _ in eligible}):
                st.markdown(f"**Entity {entity}**")
                for e, rm in [c for c in eligible if c[0] == entity]:
                    chk_key = f"ro_chk_{e}_{rm}"
                    tgt_key = f"ro_target_{e}_{rm}"
                    default_tgt = next_month_of(rm)

                    row = st.columns([0.06, 2.2, 0.15, 1.6])
                    with row[0]:
                        checked = st.checkbox(
                            "##", value=st.session_state.get(chk_key, False),
                            key=chk_key, label_visibility="collapsed",
                        )
                    with row[1]:
                        ready = rollable_users(e, rm, default_tgt)
                        msg = (
                            f"{len(ready)} ready" if ready else "nothing to roll"
                        )
                        color = "#2D6A2D" if ready else "#999"
                        st.markdown(
                            f'<span style="font-size:15px;">{fmt_month(rm)}</span>'
                            f'&nbsp;&nbsp;<span style="font-size:13px;color:{color};">{msg}</span>',
                            unsafe_allow_html=True,
                        )
                    if checked:
                        with row[2]:
                            st.markdown('<span style="font-size:18px;">→</span>', unsafe_allow_html=True)
                        with row[3]:
                            base = date.today()
                            opts = sorted({
                                f"{base.year + ((base.month + i - 1) // 12)}-"
                                f"{str((base.month - 1 + i) % 12 + 1).zfill(2)}"
                                for i in range(1, 14)
                            } | {default_tgt})
                            opts = [m for m in opts if m > rm]
                            cur = st.session_state.get(tgt_key, default_tgt)
                            if cur not in opts:
                                cur = default_tgt
                            tgt = st.selectbox(
                                "##", opts, index=opts.index(cur),
                                format_func=fmt_month, key=tgt_key,
                                label_visibility="collapsed",
                            )
                            actually = rollable_users(e, rm, tgt)
                            if actually:
                                selections.append((e, rm, tgt, actually))
                            else:
                                st.caption("⚠ Everyone already has data in that month")
                st.write("")

            st.divider()
            if not selections:
                st.caption("Tick a period above to plan a rollover.")
            else:
                total = sum(len(u) for _, _, _, u in selections)
                st.markdown(
                    f"**Preview** — {len(selections)} move"
                    f"{'s' if len(selections) != 1 else ''}, "
                    f"{total} report{'s' if total != 1 else ''}"
                )
                for e, rm, tgt, users in selections:
                    st.markdown(
                        f"&nbsp;&nbsp;**Entity {e}:** {fmt_month(rm)} → **{fmt_month(tgt)}**",
                        unsafe_allow_html=True,
                    )
                    for u in users:
                        rep = load_submission(u, e, rm) or {}
                        names = [
                            i.get("initiative_name", "Unnamed")
                            for i in (rep.get("initiatives") or [])
                        ]
                        st.markdown(
                            f"&nbsp;&nbsp;&nbsp;&nbsp;• **{u}** — "
                            + ", ".join(f"*{n}*" for n in names),
                            unsafe_allow_html=True,
                        )

                st.write("")
                if st.button(f"🔄 Roll forward {total} report{'s' if total != 1 else ''}",
                             type="primary", key="do_rollover"):
                    rolled = []
                    for e, rm, tgt, users in selections:
                        rolled.extend(rollover_entity(e, rm, tgt, only_users=users))
                    if rolled:
                        st.success(f"Rolled forward for: {', '.join(sorted(set(rolled)))}.")
                        st.rerun()
                    else:
                        st.warning("Nothing was rolled forward — targets already had data.")

    # ══ Archive ═══════════════════════════════════════════════════════════════
    if active_tab == "📦 Archive":
        render_archive_browser("arcadmin")

    # ══ Export ════════════════════════════════════════════════════════════════
    if active_tab == "⬇ Export":
        if not reports:
            st.info("No reports to export yet.")
        else:
            status_keys = ["in-progress", "submitted", "rejected", "approved", "archived"]

            ex_left, ex_right = st.columns([1, 1])
            with ex_left:
                st.markdown("**1. Periods and entities**")
                chk_keys = [f"ex_chk_{e}_{m}" for e, m in combos]
                _select_all_controls("ex", chk_keys, "periods")
                st.write("")
                chosen_combos = []
                for entity in sorted({e for e, _ in combos}):
                    st.markdown(f"**Entity {entity}**")
                    for e, rm in sorted([c for c in combos if c[0] == entity], reverse=True):
                        key = f"ex_chk_{e}_{rm}"
                        if st.checkbox(
                            f"  {fmt_month(rm)}",
                            value=st.session_state.get(key, True),
                            key=key,
                        ):
                            chosen_combos.append((e, rm))
                    st.write("")

            with ex_right:
                st.markdown("**2. Status**")
                st.write("")
                chosen_statuses = []
                for key in status_keys:
                    default = key in ("submitted", "approved", "archived")
                    if st.checkbox(
                        STATUS_LABELS.get(key, key),
                        value=st.session_state.get(f"ex_st_{key}", default),
                        key=f"ex_st_{key}",
                    ):
                        chosen_statuses.append(key)

                st.write("")
                group_by = st.radio(
                    "Tab layout",
                    ["One tab per entity", "One tab per team member"],
                    key="ex_group_by",
                )

            st.divider()

            combo_set = set(chosen_combos)
            matched = [
                r for r in reports
                if (r.get("entity", ""), r.get("reporting_month", "")) in combo_set
                and computed_status(r) in chosen_statuses
            ]

            if not chosen_combos or not chosen_statuses:
                st.warning("Select at least one period and one status.")
            else:
                unique_rows, dropped = dedupe_stats(matched)
                sc1, sc2, sc3, sc4 = st.columns(4)
                sc1.metric("Entities",    len({r.get("entity") for r in matched}))
                sc2.metric("Reports",     len(matched))
                sc3.metric("Team members", len({r.get("user") for r in matched}))
                sc4.metric("Rows",        unique_rows)

                if dropped:
                    # Rolling forward copies earlier rows into later reports, so
                    # selecting overlapping periods sees the same row more than
                    # once. Say so rather than quietly changing the count.
                    st.caption(
                        f"ℹ {dropped} duplicate row{'s' if dropped != 1 else ''} "
                        "collapsed — later periods carry copies of earlier months, "
                        "and the most recently reviewed version is kept."
                    )

                if not matched:
                    st.caption("No reports match this combination.")
                else:
                    gb = "user" if group_by.endswith("member") else "entity"
                    tag = ("_".join(
                        STATUS_LABELS[s].split(" ", 1)[-1].replace(" ", "")
                        for s in chosen_statuses
                    ) if len(chosen_statuses) < len(status_keys) else "AllStatuses")
                    st.download_button(
                        f"↓ Download consolidated report ({len(matched)} report"
                        f"{'s' if len(matched) != 1 else ''})",
                        data=build_excel_consolidated(
                            matched, group_by=gb,
                            status_desc=status_desc_for(chosen_statuses),
                        ),
                        file_name=f"Consolidated_Report_{tag}_{datetime.now().strftime('%m%d%y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                    )

    # ══ Backup & Data ═════════════════════════════════════════════════════════
    if active_tab == "🗄 Backup & Data":
        st.markdown("**Backup**")
        st.caption(
            "Every report, every period, every status, as one Excel file. "
            "This app's storage is wiped whenever the server restarts, so "
            "download this after each approval cycle and keep it somewhere durable."
        )
        backup_bytes = create_backup_excel()
        col1, col2 = st.columns([2, 3])
        with col1:
            st.download_button(
                "↓ Download full data backup (.xlsx)",
                data=backup_bytes,
                file_name=f"RD_Data_Backup_{datetime.now().strftime('%m%d%y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col2:
            # This count used to glob every JSON in data/ and call .get() on each,
            # which crashed with an AttributeError as soon as anyone added a custom
            # entity, because that config file holds a list rather than a dict.
            st.caption(
                f"{store.count_reports()} report"
                f"{'s' if store.count_reports() != 1 else ''} with data · "
                f"{len(backup_bytes) // 1024} KB"
            )

        st.divider()

        st.markdown("**⚠ Delete all reports**")
        st.caption(
            "Permanently deletes every report. Your entity list and permissions "
            "are stored separately and are kept. Download the backup first — "
            "this can't be undone."
        )
        if not st.session_state.get("confirm_delete_all"):
            if st.button("🗑 Delete all reports", key="del_all_btn"):
                st.session_state.confirm_delete_all = True
                st.rerun()
        else:
            st.error(
                f"This deletes all {store.count_reports()} reports for every team "
                "member across every period. Entity list and permissions are kept."
            )
            da1, da2 = st.columns(2)
            with da1:
                if st.button("Yes, delete all reports", key="del_all_confirm", type="primary"):
                    n = delete_all_history()
                    st.session_state.confirm_delete_all = False
                    st.success(f"Deleted {n} reports.")
                    st.rerun()
            with da2:
                if st.button("Cancel", key="del_all_cancel"):
                    st.session_state.confirm_delete_all = False
                    st.rerun()

        st.divider()

        st.markdown("**📤 Restore from backup**")
        st.caption(
            "If the server restarted and lost its data, upload a backup or export "
            "here to rebuild the reports from it. Works with the full backup, a "
            "consolidated report, or a single person's export."
        )

        uploaded = st.file_uploader(
            "Upload a backup or export file", type=["xlsx"], key="restore_uploader"
        )

        if uploaded is not None:
            parsed, parse_warnings = parse_import_workbook(uploaded.getvalue())

            if not parsed:
                st.error("No usable report data found in this file.")
                for w in parse_warnings:
                    st.caption(f"⚠ {w}")
            else:
                will_import, will_skip = [], []
                for (username, entity, rm), data in parsed.items():
                    existing = load_submission(username, entity, rm)
                    row = {
                        "User": username, "Entity": entity, "Period": fmt_month(rm),
                        "Status": STATUS_LABELS.get(data["status"], data["status"]),
                        "Initiatives": len(data["initiatives"]),
                    }
                    (will_skip if (existing and existing.get("initiatives")) else will_import).append(row)

                pc1, pc2, pc3 = st.columns(3)
                pc1.metric("Found", len(parsed))
                pc2.metric("Will import", len(will_import))
                pc3.metric("Already present", len(will_skip))

                if will_import:
                    st.markdown("**Will be imported:**")
                    for row in will_import:
                        st.markdown(
                            f"&nbsp;&nbsp;• **{row['User']}** — Entity {row['Entity']} — "
                            f"{row['Period']} — {row['Status']} "
                            f"({row['Initiatives']} initiative"
                            f"{'s' if row['Initiatives'] != 1 else ''})",
                            unsafe_allow_html=True,
                        )

                if will_skip:
                    with st.expander(f"⚠ {len(will_skip)} already have data — skipped unless you overwrite"):
                        for row in will_skip:
                            st.markdown(
                                f"&nbsp;&nbsp;• **{row['User']}** — Entity {row['Entity']} — "
                                f"{row['Period']} — {row['Status']}",
                                unsafe_allow_html=True,
                            )

                if parse_warnings:
                    with st.expander(f"ℹ {len(parse_warnings)} note(s) about this import"):
                        for w in parse_warnings:
                            st.caption(w)

                overwrite = False
                if will_skip:
                    overwrite = st.checkbox(
                        "Overwrite periods that already have data",
                        value=False, key="restore_overwrite",
                    )

                n_apply = len(parsed) if overwrite else len(will_import)
                if n_apply == 0:
                    st.info("Nothing to import — every period in this file already has data.")
                elif st.button(f"📤 Restore {n_apply} report{'s' if n_apply != 1 else ''}",
                               type="primary", key="confirm_restore"):
                    result = apply_import(parsed, overwrite=overwrite)
                    st.success(
                        f"Restored {len(result['imported'])} report(s)."
                        + (f" Skipped {len(result['skipped'])}." if result["skipped"] else "")
                    )
                    st.rerun()

    # ══ Settings ══════════════════════════════════════════════════════════════
    if active_tab == "⚙ Settings":
        st.markdown("**Entity filing permissions**")
        st.caption(
            "Controls which entities each person can FILE reports for. This is not "
            "a confidentiality setting — the Archive tab shows archived reports for "
            "every entity to everyone, by design."
        )

        all_ents = all_entities()
        perms = load_permissions()
        employees_list = [e for e in EMPLOYEES if e != "Admin"]

        changed = False
        for emp in employees_list:
            current = perms.get(emp, [])
            perm_key = f"perm_{emp}"
            if perm_key not in st.session_state:
                st.session_state[perm_key] = current if current else all_ents
            col1, col2 = st.columns([2, 4])
            with col1:
                st.markdown(f"**{emp}**")
            with col2:
                sel = st.multiselect(
                    f"##perm_{emp}", all_ents, key=perm_key,
                    label_visibility="collapsed",
                    placeholder="All entities (no restriction)",
                )
                new_perm = [] if set(sel) == set(all_ents) else list(sel)
                if new_perm != current:
                    perms[emp] = new_perm
                    changed = True

        st.write("")
        if changed:
            save_permissions(perms)
            st.success("Permissions saved.")

        st.divider()
        st.markdown("**Current restrictions**")
        restricted = {u: v for u, v in perms.items() if v}
        if not restricted:
            st.caption("No restrictions — everyone can file for all entities.")
        else:
            for u, ents in restricted.items():
                st.markdown(f"&nbsp;&nbsp;• **{u}**: {', '.join(ents)}", unsafe_allow_html=True)

        st.write("")
        if st.button("Reset all to all entities", key="reset_perms"):
            save_permissions({})
            for emp in employees_list:
                st.session_state.pop(f"perm_{emp}", None)
            st.rerun()

        st.divider()
        st.markdown("**Entities**")
        st.caption(", ".join(all_ents))

        st.divider()
        render_closed_bc_settings()

        st.divider()
        render_directory_settings()


# ── Closed business components admin ──────────────────────────────────────────

def render_closed_bc_settings():
    """Review and reopen closed business components.

    Closure is deliberately one-way from the preparer's side: they can close a
    component when its work finishes, but only the Oversight Lead can reopen
    one, since reopening means work the client already reported as finished has
    started again.
    """
    st.markdown("**Closed business components**")
    st.caption(
        "Closed when a preparer finishes the last of the work under a component. "
        "A closed component can't be used on new initiatives — reopen it here if "
        "work restarts. This is separate from resolving a technical uncertainty, "
        "which closes a question rather than the component."
    )

    ents = all_entities()
    any_closed = False
    for entity in ents:
        recs = closed_bc_list(entity)
        if not recs:
            continue
        any_closed = True
        st.markdown(f"*Entity {entity}*")
        for rec in recs:
            label = rec.get("label", "—")
            when  = ts_to_et(rec.get("closed_at"), "%b %d, %Y") if rec.get("closed_at") else ""
            cc1, cc2 = st.columns([5, 1])
            with cc1:
                st.markdown(
                    f"&nbsp;&nbsp;• **{label}**  \n"
                    f"<small style='color:#64748b;'>"
                    f"closed by {rec.get('closed_by','—')}"
                    + (f" · {when}" if when else "")
                    + (f" · {rec.get('note')}" if rec.get("note") else "")
                    + "</small>",
                    unsafe_allow_html=True,
                )
            with cc2:
                key = f"reopen_bc_{entity}_{_safe_name(label)}"
                if st.button("↩ Reopen", key=key):
                    reopen_business_component(entity, label)
                    st.success(f"Reopened {label} for Entity {entity}.")
                    st.rerun()
        st.write("")

    if not any_closed:
        st.caption("Nothing closed yet.")


# ── Employee directory admin ──────────────────────────────────────────────────

def render_directory_settings():
    """Upload / replace the client's employee list.

    Deliberately tolerant about file shape: HR exports arrive with title rows,
    notes, and merged cells above the real header — the current client's has
    five — and asking the client to clean that up is a worse ask than detecting
    it here. The admin confirms the sheet and the two columns that matter;
    everything else in the file is ignored.
    """
    st.markdown("**Employee directory**")

    if directory.is_loaded():
        m = directory.meta()
        n_emp, n_grp = directory.counts()
        when = ts_to_et(m.get("saved_at"), "%b %d, %Y %I:%M %p") if m.get("saved_at") else ""
        st.success(
            f"✓ **{n_emp:,} employees** across **{n_grp} {GROUP_LABEL.lower()}s** — "
            f"from *{m.get('filename','uploaded file')}*"
            + (f", sheet *{m.get('sheet','')}*" if m.get("sheet") else "")
            + (f", loaded {when}" if when else "")
        )
        st.caption(
            f"Preparers pick a {GROUP_LABEL.lower()} first, then choose team "
            "members from within it. Uploading again replaces this; initiatives "
            "already saved keep the names they were given either way."
        )
        if st.session_state.get("dir_confirm_clear"):
            st.warning(
                f"Remove the directory? {GROUP_LABEL} becomes free text and team "
                "selection falls back to the built-in list. Saved initiatives "
                "are not affected."
            )
            dc1, dc2 = st.columns(2)
            with dc1:
                if st.button("Yes, remove it", key="dir_clear_yes", type="primary"):
                    directory.clear()
                    st.session_state.dir_confirm_clear = False
                    st.rerun()
            with dc2:
                if st.button("Cancel", key="dir_clear_no"):
                    st.session_state.dir_confirm_clear = False
                    st.rerun()
        else:
            if st.button("Remove directory", key="dir_clear"):
                st.session_state.dir_confirm_clear = True
                st.rerun()
    else:
        st.info(
            f"No directory loaded, so {GROUP_LABEL} is free text and team "
            f"selection uses the built-in list of {len(EMPLOYEES)} names."
        )

    st.caption(
        "⚠ This app's storage is wiped when the server restarts, so the "
        "directory has to be re-uploaded afterwards along with the reports."
    )

    st.write("")
    up = st.file_uploader(
        "Upload an employee list (.xlsx)", type=["xlsx"], key="dir_uploader"
    )
    if up is None:
        return

    try:
        raw = up.getvalue()
        sheets = directory.sheet_names(raw)
    except Exception as e:
        st.error(f"Couldn't open that file as an Excel workbook: {e}")
        return
    if not sheets:
        st.error("That workbook has no sheets.")
        return

    sheet = st.selectbox("Sheet", sheets, key="dir_sheet")
    try:
        info = directory.inspect(raw, sheet)
    except Exception as e:
        st.error(f"Couldn't read that sheet: {e}")
        return
    if not info["columns"]:
        st.error("No column headers found on that sheet — try another one.")
        return

    st.caption(
        f"Header detected on row **{info['header_row']}** · "
        f"**{info['n_rows']:,}** data rows · **{len(info['columns'])}** columns"
    )

    cols, sug = info["columns"], info["suggested"]

    def _idx(val, fallback=0):
        return cols.index(val) if val in cols else fallback

    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        name_col = st.selectbox(
            "Name column", cols, index=_idx(sug.get("name")), key="dir_name_col"
        )
    with mc2:
        group_col = st.selectbox(
            f"{GROUP_LABEL} column", cols, index=_idx(sug.get("group")),
            key="dir_group_col",
            help="Usually Cost Center. Whatever you pick here is what preparers "
                 "filter employees by.",
        )
    with mc3:
        id_opts = ["(none)"] + cols
        id_col = st.selectbox(
            "Employee ID (optional)", id_opts,
            index=(id_opts.index(sug["id"]) if sug.get("id") in id_opts else 0),
            key="dir_id_col",
            help="Only used to tell apart two people with the same name.",
        )

    if name_col == group_col:
        st.error(f"Name and {GROUP_LABEL.lower()} must be different columns.")
        return

    try:
        records, warnings = directory.parse(
            raw, sheet, name_col, group_col, "" if id_col == "(none)" else id_col
        )
    except Exception as e:
        st.error(f"Couldn't read that sheet: {e}")
        return
    if not records:
        for w in warnings:
            st.error(w)
        return

    sizes_by_group = {}
    for r in records:
        sizes_by_group[r["group"]] = sizes_by_group.get(r["group"], 0) + 1
    sizes = sorted(sizes_by_group.values(), reverse=True)

    pc1, pc2, pc3 = st.columns(3)
    pc1.metric("Employees", f"{len(records):,}")
    pc2.metric(f"{GROUP_LABEL}s", len(sizes_by_group))
    pc3.metric("Largest group", sizes[0] if sizes else 0)

    for w in warnings:
        st.caption(f"ℹ {w}")

    if sizes and sizes[0] > 40:
        n_big = sum(1 for s in sizes if s > 40)
        st.caption(
            f"ℹ {n_big} group{'s' if n_big != 1 else ''} "
            f"{'have' if n_big != 1 else 'has'} more than 40 people. Those lists "
            "are long — the picker's search box handles them, but it's worth "
            "knowing before the client sees it."
        )

    with st.expander("Preview the first 10 rows"):
        st.dataframe(
            [{"Name": r["display"], GROUP_LABEL: r["group"]} for r in records[:10]],
            width='stretch', hide_index=True,
        )

    st.write("")
    if st.button(
        f"Load {len(records):,} employees"
        + (" (replaces the current directory)" if directory.is_loaded() else ""),
        type="primary", key="dir_save",
    ):
        directory.save(records, {
            "filename": up.name, "sheet": sheet,
            "name_col": name_col, "group_col": group_col,
        })
        st.success(
            f"Loaded {len(records):,} employees across "
            f"{len(sizes_by_group)} {GROUP_LABEL.lower()}s."
        )
        st.rerun()


# ── Entry Picker Screen ───────────────────────────────────────────────────────

def screen_entry_picker():
    """
    Shown after clicking + Add Initiative.
    Lets the user choose between Guided Entry (wizard) or Bulk Entry (grid).
    """
    draft  = st.session_state.draft
    entity = draft.get("entity", "")
    rm     = draft.get("reporting_month", "")
    am     = draft.get("activities_month", "") or rm

    st.markdown("## ＋ Add Initiative")
    st.caption(
        f"Entity **{entity}** · Filing: **{fmt_month(rm)}** · "
        f"Reporting Period: **{fmt_month(am)}**"
    )
    st.divider()
    st.markdown("### How would you like to enter your initiative(s)?")
    st.write("")

    c1, c2 = st.columns(2)

    with c1:
        st.markdown("#### 📝 Guided Entry")
        st.write(
            "Step-by-step wizard that walks you through each field one at a time. "
            "Best for adding one or two initiatives with full detail."
        )
        st.write("")
        if st.button("Start Guided Entry →", type="primary", width='stretch', key="pick_guided"):
            st.session_state.wiz_init = new_initiative()
            st.session_state.wiz_step = 0
            st.session_state.wiz_mode = "new"
            st.session_state.screen   = "wizard"
            st.rerun()

    with c2:
        st.markdown("#### 📊 Bulk Entry")
        st.write(
            "Spreadsheet-style grid where you fill in multiple rows at once. "
            "Best for entering several initiatives in one sitting."
        )
        st.write("")
        if st.button("Start Bulk Entry →", width='stretch', key="pick_bulk"):
            st.session_state.pop("bulk_df", None)
            st.session_state.screen = "bulk_entry"
            st.rerun()

    st.write("")
    st.divider()
    if st.button("← Back to Dashboard", key="pick_back"):
        st.session_state.screen = "dashboard"
        st.rerun()


# ── Bulk Entry Screen ──────────────────────────────────────────────────────────

def screen_bulk_entry():
    """
    Spreadsheet-style bulk initiative entry.
    Each row = one initiative. Adds to the existing draft on save.

    Wrapped in st.form so that cell edits do NOT trigger reruns —
    the grid stays stable while the user fills in multiple rows quickly.
    Only the Save / Back buttons cause a rerun.
    """
    import pandas as pd

    user  = st.session_state.user
    draft = st.session_state.draft
    entity = draft.get("entity", "")
    rm     = draft.get("reporting_month", "")
    am     = draft.get("activities_month", "") or rm

    st.markdown("## 📊 Bulk Initiative Entry")
    st.caption(
        f"Entity **{entity}** · Filing: **{fmt_month(rm)}** · Reporting Period: **{fmt_month(am)}**  \n"
        "Fill in each row — one initiative per row. Required fields are marked \\*. "
        "Completely blank rows are skipped automatically."
    )
    st.divider()

    if "bulk_df" not in st.session_state:
        n = 5
        import pandas as _pd
        st.session_state.bulk_df = pd.DataFrame({
            f"{GROUP_LABEL} *":    [""] * n,
            "Initiative Name *":   [""] * n,
            "Business Component *":[""] * n,
            "Description *":       [""] * n,
            "Tech Uncertainty *":  [""] * n,
            "Start Date *":        [_pd.NaT] * n,
            "End Date *":          [_pd.NaT] * n,
            "Activities *":        [""] * n,
            "Notes":               [""] * n,
        }, index=range(1, n + 1))

    st.caption(
        "💡 **Tips:** Fill any number of rows freely — the grid won't reload while you type. "
        "Tab between cells · Date cells have a calendar picker · "
        "Team Member cells have a dropdown · Use the ＋ row button to add more rows"
    )

    # Wrap in st.form — CRITICAL fix: prevents a rerun on every cell edit.
    # Without this, each Tab/click triggers a full page reload which loses
    # in-progress edits when filling multiple rows quickly.
    with st.form("bulk_entry_form", clear_on_submit=False):
        edited = st.data_editor(
            st.session_state.bulk_df,
            num_rows="dynamic",
            width='stretch',
            hide_index=False,
            column_config={
                # Rows can each have a different group — the team-member step
                # that follows scopes each row's people to its own group.
                f"{GROUP_LABEL} *": (
                    st.column_config.SelectboxColumn(
                        f"{GROUP_LABEL} *", width="medium",
                        options=directory.groups(),
                        help="Team members for this row are limited to this group.",
                    )
                    if directory.is_loaded() else
                    st.column_config.TextColumn(
                        f"{GROUP_LABEL} *", width="medium",
                        help="No employee directory loaded — free text for now.",
                    )
                ),
                "Initiative Name *": st.column_config.TextColumn(
                    "Initiative Name *", width="medium",
                    help="Short, unique name for this initiative",
                ),
                "Business Component *": st.column_config.TextColumn(
                    "Business Component *", width="medium",
                    help="The business component this R&D work belongs to",
                ),
                "Description *": st.column_config.TextColumn(
                    "Initiative Description *", width="large",
                    help="What this initiative is trying to accomplish",
                ),
                "Tech Uncertainty *": st.column_config.TextColumn(
                    "Technical Uncertainty *", width="large",
                    help="What technical question or uncertainty you are trying to resolve",
                ),
                "Start Date *": st.column_config.DateColumn(
                    "Start Date *", format="YYYY-MM-DD",
                    help="When work on this initiative began",
                ),
                "End Date *": st.column_config.DateColumn(
                    "Expected End Date *", format="YYYY-MM-DD",
                    help="Expected completion date",
                ),
                "Activities *": st.column_config.TextColumn(
                    "Activities *", width="large",
                    help="Activities this month to eliminate the technical uncertainty",
                ),

                "Notes": st.column_config.TextColumn(
                    "Notes", width="medium",
                    help="Optional — any additional context, blockers, or upcoming steps",
                ),
            },
        )

        st.write("")
        fb1, fb2, fb3 = st.columns([1, 1, 1])
        with fb1:
            go_back = st.form_submit_button("← Back to Dashboard")
        with fb3:
            do_next = st.form_submit_button("Next → Add Team Members", type="primary")

    # Handle buttons outside the form context
    if go_back:
        st.session_state.screen = "dashboard"
        st.rerun()

    if do_next:
        st.session_state.bulk_df = edited

        errors, to_add = [], []

        for idx, row in edited.iterrows():
            row_num = idx if isinstance(idx, int) else int(idx)

            grp   = str(row.get(f"{GROUP_LABEL} *",   "") or "").strip()
            name  = str(row.get("Initiative Name *",  "") or "").strip()
            bc    = str(row.get("Business Component *","") or "").strip()
            desc  = str(row.get("Description *",      "") or "").strip()
            unc   = str(row.get("Tech Uncertainty *", "") or "").strip()
            acts  = str(row.get("Activities *",       "") or "").strip()
            notes = str(row.get("Notes",              "") or "").strip()
            import pandas as _pd
            sd_raw = row.get("Start Date *")
            ed_raw = row.get("End Date *")
            # DateColumn returns date objects or pd.NaT — normalise to str or None
            def _to_str(v):
                if v is None or (hasattr(v, "__class__") and v.__class__.__name__ == "NaTType"):
                    return ""
                try:
                    import pandas as __pd
                    if __pd.isnull(v):
                        return ""
                except Exception:
                    pass
                if hasattr(v, "strftime"):
                    return v.strftime("%Y-%m-%d")
                return str(v).strip()
            sd = _to_str(sd_raw)
            ed = _to_str(ed_raw)

            if not any([grp, name, bc, desc, unc, acts]) and not sd and not ed:
                continue

            row_errors = []
            if bc and bc_is_closed(draft.get("entity", ""), bc):
                errors.append(
                    f'Row {row_num} — business component "{bc}" is closed and '
                    "can't be reused. Pick another, or ask the Oversight Lead to "
                    "reopen it."
                )
                continue
            if not grp and directory.is_loaded():
                row_errors.append(GROUP_LABEL)
            if not name:   row_errors.append("Initiative Name")
            if not bc:     row_errors.append("Business Component")
            if not desc:   row_errors.append("Description")
            if not unc:    row_errors.append("Tech Uncertainty")
            if not acts:   row_errors.append("Activities")
            if not sd:     row_errors.append("Start Date")
            if not ed:     row_errors.append("End Date")

            if row_errors:
                errors.append(f"Row {row_num} — missing: {', '.join(row_errors)}")
                continue

            sd_str, ed_str = sd, ed

            existing_names = [i.get("initiative_name","").strip().lower()
                              for i in draft.get("initiatives", [])]
            batch_names    = [r.get("initiative_name","").strip().lower() for r in to_add]
            if name.lower() in existing_names or name.lower() in batch_names:
                errors.append(f'Row {row_num} — "{name}" already exists in this report.')
                continue

            init = new_initiative()
            init["employee_group"]         = grp
            init["initiative_name"]        = name
            init["business_component"]     = bc
            init["initiative_description"] = desc
            init["tech_uncertainty"]       = unc
            init["start_date"]             = sd_str
            init["expected_end_date"]      = ed_str
            init["activities"]             = acts
            init["team_members"]           = []
            init["notes"]                  = notes
            init["month_yr"]               = am
            to_add.append(init)

        if errors:
            for e in errors:
                st.error(e)
        elif not to_add:
            st.warning("No initiatives to save — fill in at least one row.")
        else:
            st.session_state.bulk_pending = to_add
            st.session_state.screen = "bulk_team"
            st.rerun()


# ── Bulk Team Member Assignment ───────────────────────────────────────────────

def screen_bulk_team():
    """
    Step 2 of bulk entry — assign team members to each validated initiative.
    Shows initiative name and reporting period so the user knows exactly
    which one they're adding members to.
    """
    user  = st.session_state.user
    draft = st.session_state.draft
    inits = st.session_state.get("bulk_pending", [])
    am    = draft.get("activities_month", "") or draft.get("reporting_month", "")

    if not inits:
        st.session_state.screen = "dashboard"
        st.rerun()

    st.markdown("## 👥 Add Team Members")
    st.caption(
        f"Reporting Period: **{fmt_month(am)}**  \n"
        "Select team members for each initiative. You can leave any blank and add them later via Edit."
    )
    st.divider()

    # Collect selections — use seeded session state keys so they persist
    # if the user goes back and forward
    for i, init in enumerate(inits):
        iname = init.get("initiative_name", f"Initiative {i+1}")
        bc    = init.get("business_component", "")
        sd    = init.get("start_date", "—")
        ed    = init.get("expected_end_date", "—")

        grp = init.get("employee_group", "")
        st.markdown(f"**{iname}** — *{bc}*")
        st.caption(
            f"📅 {sd} → {ed}  ·  Reporting Period: {fmt_month(am)}"
            + (f"  ·  {GROUP_LABEL}: **{grp}**" if grp else "")
        )

        split_legacy_members(init)
        render_team_picker(
            f"bulk_team_{i}", grp, init.get("team_members", []),
            current_contractors=init.get("contractors") or [],
        )
        st.divider()

    st.divider()

    b1, b2, b3 = st.columns([1, 1, 1])
    with b1:
        if st.button("← Back to Grid", key="bt_back"):
            st.session_state.screen = "bulk_entry"
            st.rerun()
    with b3:
        if st.button("Save All Initiatives ✓", type="primary", key="bt_save"):
            # Assign team members from each multiselect into the initiatives
            for i, init in enumerate(inits):
                init["team_members"] = st.session_state.get(f"bulk_team_{i}", [])
                init["contractors"]  = st.session_state.get(f"bulk_team_{i}__con", [])

            # Append all to the draft and save
            draft["initiatives"] = draft.get("initiatives", []) + inits
            save_draft(user, draft)
            st.session_state.draft = draft

            # Clean up
            st.session_state.pop("bulk_pending", None)
            st.session_state.pop("bulk_df", None)
            for i in range(len(inits)):
                clear_team_picker_state(f"bulk_team_{i}")
                st.session_state.pop(f"bulk_team_{i}", None)

            n = len(inits)
            st.success(f"✓ Saved {n} initiative{'s' if n!=1 else ''}.")
            st.session_state.screen = "dashboard"
            st.rerun()


# ── Pathway Screens ────────────────────────────────────────────────────────────

def screen_pathway_select():
    """
    Shown before the wizard when a user edits a carry-over initiative.
    Asks: what happened with this initiative this month?
    Routes to the appropriate lightweight or full form.
    """
    user  = st.session_state.user
    init  = st.session_state.wiz_init
    draft = st.session_state.draft

    st.markdown("## What happened with this initiative this month?")
    st.markdown(
        f"**{init.get('initiative_name', 'Unnamed')}** — *{init.get('business_component', '')}*"
    )
    st.divider()

    existing_pathway = init.get("pathway", "")
    if existing_pathway:
        pathway_labels = {"continuing": "Still working on it",
                          "resolved": "Resolved this month",
                          "new_direction": "New direction"}
        st.caption(f"Previously chose: **{pathway_labels.get(existing_pathway, existing_pathway)}** — select again to change.")
        st.write("")

    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("### 🔄 Still working on it")
        st.caption("Lightweight monthly update. Update activities and notes; all other details carry over as-is.")
        if st.button("Select →", key="pw_continuing", width='stretch', type="primary"):
            st.session_state.wiz_init["pathway"] = "continuing"
            st.session_state.screen = "continuing"
            st.rerun()

    with c2:
        st.markdown("### ✅ Resolved this month")
        st.caption("Capture the resolution and outcome. Sets a completion date; won't appear in future carry-overs.")
        if st.button("Select →", key="pw_resolved", width='stretch'):
            st.session_state.wiz_init["pathway"] = "resolved"
            st.session_state.screen = "resolved"
            st.rerun()

    with c3:
        st.markdown("### 🔀 New direction")
        st.caption("The scope or approach has fundamentally changed. Full structured entry — same as adding a new initiative.")
        if st.button("Select →", key="pw_new_direction", width='stretch'):
            st.session_state.wiz_init["pathway"] = "new_direction"
            st.session_state.wiz_mode = "edit"
            st.session_state.wiz_step = 0
            st.session_state.screen = "wizard"
            st.rerun()

    st.write("")
    st.caption(
        "Every initiative carried into this period needs one of these three "
        "before the report is submitted."
    )
    if st.button("⏳ Skip for now", key="pw_back",
                 help="Come back to it later in this session"):
        # Recorded so the dashboard doesn't immediately send us straight back
        # here — without this the auto-prompt and the back button loop.
        skipped = set(st.session_state.get("pathway_skipped") or set())
        skipped.add(init.get("id"))
        st.session_state.pathway_skipped = skipped
        st.session_state.screen = "dashboard"
        st.rerun()


def _save_initiative_update(user: str, init: dict):
    """Save updated initiative back into the draft and return to dashboard."""
    draft = st.session_state.draft
    if not init.get("month_yr"):
        init["month_yr"] = draft.get("activities_month") or draft.get("reporting_month", "")
    # The user has just revised this initiative, so it is no longer awaiting
    # revision. The reviewer's note moves into history rather than continuing to
    # display as an open problem on work that has since been fixed.
    clear_return_flag(init)
    draft["initiatives"] = [
        i if i["id"] != init["id"] else init
        for i in draft["initiatives"]
    ]
    save_draft(user, draft)
    st.session_state.draft  = draft
    st.session_state.screen = "dashboard"
    st.rerun()


def screen_continuing():
    """Pathway 1 — Still Working On It. Prominent activities, dates + notes; other fields collapsible."""
    user  = st.session_state.user
    init  = dict(st.session_state.wiz_init)

    # The picker seeds itself from the initiative. This used to seed with
    # `if m in EMPLOYEES`, which silently dropped every team member who wasn't
    # in the built-in ten — all directory names, and all write-ins.

    st.markdown(f"## 🔄 Monthly Update")
    st.markdown(f"**{init.get('initiative_name', 'Unnamed')}** — *{init.get('business_component', '')}*")
    st.caption("Update what changed this month. All fields are pre-filled from last month.")
    st.divider()

    activities = st.text_area(
        "What did you work on this month? *",
        value=init.get("activities", ""),
        height=160,
        placeholder="Describe the activities you carried out this month to advance this initiative...",
        key="cont_activities",
    )

    dc1, dc2 = st.columns(2)
    with dc1:
        sd = st.text_input(
            "Start Date (YYYY-MM-DD)",
            value=init.get("start_date") or "",
            key="cont_sd",
        )
    with dc2:
        ed = st.text_input(
            "Expected End Date (YYYY-MM-DD)",
            value=init.get("expected_end_date") or "",
            key="cont_ed",
        )

    notes = st.text_area(
        "Notes (optional)",
        value=init.get("notes", ""),
        height=80,
        placeholder="Any blockers, upcoming steps, or context worth noting...",
        key="cont_notes",
    )

    st.write("")

    with st.expander("✏ Update other details (optional)", expanded=False):
        st.caption("These fields carried over from last month. Update if anything has changed.")
        bc   = st.text_input("Business Component", value=init.get("business_component", ""), key="cont_bc")
        name = st.text_input("Initiative Name",    value=init.get("initiative_name", ""),    key="cont_name")
        desc = st.text_area("Initiative Description", value=init.get("initiative_description", ""), height=100, key="cont_desc")
        unc  = st.text_area("Technical Uncertainty",  value=init.get("tech_uncertainty", ""),        height=100, key="cont_unc")
        new_grp = group_select("cont_grp", init.get("employee_group", ""))
        if new_grp and new_grp != init.get("employee_group", ""):
            init["employee_group"] = new_grp
            kept = _prune_to_group(init.get("team_members"), new_grp)
            if len(kept) != len(init.get("team_members") or []):
                st.warning(
                    f"Team members outside {new_grp} were removed. "
                    "Reselect below."
                )
                init["team_members"] = kept
                clear_team_picker_state("cont_team")
                st.session_state.pop("cont_team", None)
        split_legacy_members(init)
        team, cont_cons = render_team_picker(
            "cont_team", init.get("employee_group", ""),
            init.get("team_members") or [],
            current_contractors=init.get("contractors") or [],
        )

    st.write("")
    b1, b2 = st.columns([1, 1])
    with b1:
        if st.button("← Back", key="cont_back"):
            st.session_state.screen = "pathway_select"
            st.rerun()
    with b2:
        if st.button("Save Update ✓", type="primary", key="cont_save"):
            if not activities.strip():
                st.error("Please describe what you worked on this month.")
            else:
                init["activities"]             = activities.strip()
                init["notes"]                  = notes.strip()
                init["start_date"]             = st.session_state.get("cont_sd") or None
                init["expected_end_date"]      = st.session_state.get("cont_ed") or None
                init["business_component"]     = st.session_state.get("cont_bc", init["business_component"])
                init["initiative_name"]        = st.session_state.get("cont_name", init["initiative_name"])
                init["initiative_description"] = st.session_state.get("cont_desc", init["initiative_description"])
                init["tech_uncertainty"]       = st.session_state.get("cont_unc", init["tech_uncertainty"])
                init["team_members"]           = st.session_state.get("cont_team", init.get("team_members", []))
                init["contractors"]            = st.session_state.get("cont_team__con", init.get("contractors", []))
                init["employee_group"]         = st.session_state.get(
                    "cont_grp__grp", init.get("employee_group", "")
                ) if directory.is_loaded() else init.get("employee_group", "")
                _save_initiative_update(user, init)


def screen_resolved():
    """Pathway 2 — Resolved This Month. Captures final activities, dates, completion date, and outcome."""
    user  = st.session_state.user
    init  = dict(st.session_state.wiz_init)

    bc     = (init.get("business_component") or "").strip()
    entity = st.session_state.draft.get("entity", "")

    st.markdown(f"## ✅ Resolution — {init.get('initiative_name', 'Unnamed')}")
    st.markdown(f"*{bc}*")
    st.caption(
        "Capture what you did to reach resolution and the outcome. This "
        "initiative won't appear in next month's carry-over suggestions."
    )
    st.divider()

    activities = st.text_area(
        "What activities led to resolution this month? *",
        value=init.get("activities", ""),
        height=140,
        placeholder="Describe the final activities that eliminated the technical uncertainty...",
        key="res_activities",
    )

    rd1, rd2 = st.columns(2)
    with rd1:
        sd = st.text_input(
            "Start Date (YYYY-MM-DD)",
            value=init.get("start_date") or "",
            key="res_sd",
        )
    with rd2:
        ed = st.text_input(
            "Expected End Date (YYYY-MM-DD)",
            value=init.get("expected_end_date") or "",
            key="res_ed",
        )

    today_str = date.today().strftime("%Y-%m-%d")
    completion_date = st.text_input(
        "Completion Date (YYYY-MM-DD) *",
        value=init.get("completion_date") or today_str,
        key="res_completion",
    )
    outcome = st.text_area(
        "Outcome / Notes *",
        value=init.get("notes", ""),
        height=120,
        placeholder="Describe the outcome — what was resolved, what was learned, what the result was...",
        key="res_outcome",
    )

    # ── Closing the business component is a separate decision ───────────────
    # Resolving an uncertainty answers one question. A business component can
    # carry several, so it keeps running until someone says otherwise — and it
    # is asked here rather than inferred, because inferring it would silently
    # close components that still have live work under them.
    st.write("")
    st.markdown("**Does this also close the business component?**")
    already = bc_is_closed(entity, bc) if bc else None
    if not bc:
        st.caption("No business component on this initiative, so there's nothing to close.")
        close_bc_now = False
    elif already:
        when = ts_to_et(already.get("closed_at"), "%b %d, %Y") if already.get("closed_at") else ""
        st.info(
            f"**{already.get('label', bc)}** is already closed"
            + (f" (by {already.get('closed_by')}{', ' + when if when else ''})" if already.get("closed_by") else "")
            + "."
        )
        close_bc_now = False
    else:
        close_bc_now = st.checkbox(
            f"Yes — close **{bc}**. All work under it is finished.",
            value=False, key="res_close_bc",
        )
        st.caption(
            "Leave unticked if more R&D is expected under this component — "
            "resolving this uncertainty doesn't require closing it. Once closed, "
            "the component can't be used on new initiatives until an admin "
            "reopens it."
        )

    st.write("")
    b1, b2 = st.columns([1, 1])
    with b1:
        if st.button("← Back", key="res_back"):
            st.session_state.screen = "pathway_select"
            st.rerun()
    with b2:
        if st.button("Save Resolution ✓", type="primary", key="res_save"):
            errors = []
            if not activities.strip():
                errors.append("Please describe the activities that led to resolution.")
            if not completion_date.strip():
                errors.append("Please enter a completion date.")
            else:
                try:
                    datetime.strptime(completion_date.strip(), "%Y-%m-%d")
                except ValueError:
                    errors.append("Completion date must be in YYYY-MM-DD format.")
            if not outcome.strip():
                errors.append("Please describe the outcome.")
            if errors:
                for e in errors:
                    st.error(e)
            else:
                init["activities"]      = activities.strip()
                init["notes"]           = outcome.strip()
                init["start_date"]      = st.session_state.get("res_sd") or None
                init["expected_end_date"] = st.session_state.get("res_ed") or None
                init["completion_date"] = completion_date.strip()
                init["resolved"]        = True   # the technical uncertainty
                if close_bc_now and bc:
                    close_business_component(
                        entity, bc, user,
                        note=f"Closed with initiative '{init.get('initiative_name','')}'",
                    )
                    st.success(f"Business component **{bc}** closed.")
                _save_initiative_update(user, init)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    inject_css()
    init_session()
    screen = st.session_state.screen
    if screen == "login":
        screen_login()
    elif screen == "dashboard":
        screen_dashboard()
    elif screen == "wizard":
        screen_wizard()
    elif screen == "bulk_entry":
        screen_bulk_entry()
    elif screen == "bulk_team":
        screen_bulk_team()
    elif screen == "entry_picker":
        screen_entry_picker()
    elif screen == "pathway_select":
        screen_pathway_select()
    elif screen == "continuing":
        screen_continuing()
    elif screen == "resolved":
        screen_resolved()
    elif screen == "archive":
        screen_archive()
    elif screen == "admin":
        screen_admin()
    else:
        # Unknown screen name (e.g. a stale value after a code change) — send the
        # user somewhere real rather than rendering a blank page.
        st.session_state.screen = "login"
        screen_login()

if __name__ == "__main__":
    main()
